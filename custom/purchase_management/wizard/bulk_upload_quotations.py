from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from markupsafe import Markup
import base64
import io
import openpyxl
import re
from datetime import datetime

_LEGAL_SUFFIXES = re.compile(
    r'\b(s\.?a\.?s\.?|s\.?a\.?|ltda\.?|s\.?a\.?s|sas|ltda|l\.?t\.?d\.?a\.?|'
    r'inc\.?|corp\.?|llc\.?|cia\.?|co\.?|e\.?u\.?)\b',
    re.IGNORECASE
)


class PurchaseBulkUploadQuotationsWizard(models.TransientModel):
    _name = 'purchase.bulk.upload.quotations.wizard'
    _description = 'Wizard para importar cotizaciones desde Excel del proveedor'
    _transient_max_count = 100
    _transient_max_hours = 24

    """MANY2ONE"""
    request_quotation_id = fields.Many2one(
        'request.quotation', string='Solicitud de Cotización',
        readonly=True)
    """CHAR"""
    supplier_name = fields.Char(string='Nombre del proveedor',
                                help='Nombre del proveedor que envía esta cotización')
    """BINARY"""
    file_data = fields.Binary(string='Archivo Excel')
    """CHAR"""
    file_name = fields.Char(string='Nombre del archivo')
    parsed_validity_date = fields.Char(
        string='Fecha de Vigencia', readonly=True)
    parsed_currency = fields.Char(string='Moneda', readonly=True)
    parsed_trm = fields.Char(string='TRM', readonly=True)
    parsed_delivery_location = fields.Char(
        string='Lugar de Entrega', readonly=True)
    parsed_delivery_date = fields.Char(
        string='Fecha de Entrega', readonly=True)
    parsed_includes_freight = fields.Char(
        string='Incluye Flete', readonly=True)
    parsed_freight_price = fields.Char(string='Precio Flete', readonly=True)
    parsed_observations = fields.Text(string='Observaciones', readonly=True)
    parsed_products_count = fields.Integer(
        string='Cantidad de Productos', readonly=True)
    parsed_total = fields.Float(string='Total', readonly=True)
    """HTML"""
    validation_html = fields.Html(
        string='Resultado de validación', readonly=True)
    """BOOLEAN"""
    show_instructions = fields.Boolean(
        string='Mostrar instrucciones', default=True, readonly=True)
    show_supplier = fields.Boolean(
        string='Mostrar selección de proveedor', default=False, readonly=True)
    show_upload = fields.Boolean(
        string='Mostrar carga', default=False, readonly=True)
    show_validation = fields.Boolean(
        string='Mostrar validación', default=False, readonly=True)
    supplier_similarity_warning = fields.Char(readonly=True)

    def _normalize_supplier_name(self, name):
        """Strips legal suffixes and normalizes whitespace for duplicate detection."""
        normalized = _LEGAL_SUFFIXES.sub('', (name or '').lower())
        return re.sub(r'\s+', ' ', normalized).strip()

    def _find_duplicate_supplier(self, parent, name):
        """Returns the first existing quotation whose normalized name is an exact match."""
        normalized_new = self._normalize_supplier_name(name)
        existing = self.env['quotation.quotation'].search([
            (self._get_fk_field(), '=', parent.id),
        ])
        for q in existing:
            if self._normalize_supplier_name(q.supplier_name) == normalized_new:
                return q
        return None

    def _find_similar_supplier(self, parent, name):
        """Returns the first existing quotation whose normalized name contains or is contained in the new name."""
        normalized_new = self._normalize_supplier_name(name)
        existing = self.env['quotation.quotation'].search([
            (self._get_fk_field(), '=', parent.id),
        ])
        for q in existing:
            normalized_existing = self._normalize_supplier_name(q.supplier_name)
            if normalized_existing == normalized_new:
                continue
            if normalized_new in normalized_existing or normalized_existing in normalized_new:
                return q
        return None

    def _get_parent_record(self):
        """Retorna el registro padre (request.quotation)"""
        if self.request_quotation_id:
            return self.request_quotation_id
        raise UserError(_('No se encontró la solicitud asociada a esta importación.'))

    def _get_fk_field(self):
        return 'request_quotation_id'

    def action_next_to_supplier(self):
        """Navega de instrucciones a pantalla de selección de proveedor"""
        self.ensure_one()
        self.show_instructions = False
        self.show_supplier = True
        self.show_upload = False
        self.show_validation = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_next_to_upload(self):
        """Navega de selección de proveedor a pantalla de carga"""
        self.ensure_one()
        if not self.supplier_name or not self.supplier_name.strip():
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error'),
                    'message': _('Debe ingresar el nombre del proveedor para continuar.'),
                    'type': 'danger',
                    'sticky': False,
                }
            }
        parent = self._get_parent_record()
        duplicate = self._find_duplicate_supplier(parent, self.supplier_name.strip())
        if duplicate:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Proveedor duplicado'),
                    'message': _('Ya existe una cotización del proveedor "%s" para esta solicitud. Verifique el nombre e intente nuevamente.') % duplicate.supplier_name,
                    'type': 'danger',
                    'sticky': True,
                }
            }
        similar = self._find_similar_supplier(parent, self.supplier_name.strip())
        self.show_instructions = False
        self.show_supplier = False
        self.show_upload = True
        self.show_validation = False
        self.supplier_similarity_warning = (
            _('Existe un proveedor con nombre similar: "%s". Verifique que sean empresas diferentes antes de continuar.') % similar.supplier_name
        ) if similar else False
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_back_to_instructions(self):
        """Vuelve a la pantalla de instrucciones"""
        self.ensure_one()
        self.show_instructions = True
        self.show_supplier = False
        self.show_upload = False
        self.show_validation = False
        self.validation_html = False
        self.file_data = False
        self.file_name = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_back_to_supplier(self):
        """Vuelve a la pantalla de selección de proveedor"""
        self.ensure_one()
        self.show_instructions = False
        self.show_supplier = True
        self.show_upload = False
        self.show_validation = False
        self.validation_html = False
        self.file_data = False
        self.file_name = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_back_to_upload(self):
        """Vuelve a la pantalla de carga"""
        self.ensure_one()
        self.show_instructions = False
        self.show_supplier = False
        self.show_upload = True
        self.show_validation = False
        self.validation_html = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_validate_file(self):
        """Valida el archivo Excel cargado"""
        self.ensure_one()
        if not self.file_data:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error'),
                    'message': _('Debe cargar un archivo Excel para continuar.'),
                    'type': 'danger',
                    'sticky': False,
                }
            }
        try:
            data = self._analyze_quotation_excel()
            errors, warnings = self._validate_quotation_data(data)

            if errors:
                self.validation_html = self._generate_error_html(errors)
            else:
                self.validation_html = self._generate_success_html(
                    data, warnings)

            self.show_upload = False
            self.show_validation = True

            return {
                'type': 'ir.actions.act_window',
                'res_model': self._name,
                'view_mode': 'form',
                'res_id': self.id,
                'target': 'new',
            }
        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error al procesar el archivo'),
                    'message': str(e),
                    'type': 'danger',
                    'sticky': True,
                }
            }

    def _analyze_quotation_excel(self):
        """Analiza el Excel de cotización del proveedor"""
        self.ensure_one()
        file_data = base64.b64decode(self.file_data)
        file_obj = io.BytesIO(file_data)

        try:
            workbook = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(
                _('Error al leer el archivo Excel. Asegúrese de que sea un archivo .xlsx válido.'))

        data = {
            'validity_date': None,
            'currency': None,
            'trm': None,
            'delivery_location': None,
            'delivery_date': None,
            'includes_freight': None,
            'freight_price': None,
            'observations': None,
            'products': [],
            'total': 0,
        }

        # Lee la información de la cotizacion/oferta
        # Fila 11: Fecha de Vigencia (E11:G11), Moneda (J11), TRM (L11)
        data['validity_date'] = self._get_merged_value(sheet, 11, 5, 7)  # E-G
        data['currency'] = sheet.cell(row=11, column=10).value  # J
        trm_value = sheet.cell(row=11, column=12).value  # L
        data['trm'] = trm_value if trm_value != 'N/A' else None

        # Fila 12: Lugar de Entrega (E12:G12), Fecha de Entrega (J12:L12)
        data['delivery_location'] = self._get_merged_value(
            sheet, 12, 5, 7)  # E-G
        data['delivery_date'] = self._get_merged_value(
            sheet, 12, 10, 12)  # J-L

        # Fila 13: ¿Incluye Flete? (E13), Precio Est. (G13)
        data['includes_freight'] = sheet.cell(row=13, column=5).value  # E
        freight_value = sheet.cell(row=13, column=7).value  # G
        data['freight_price'] = freight_value if freight_value != 'N/A' else None

        # Busca los productos (empiezan después de la fila de encabezados "Nombre", "Cantidad", etc.)
        products_start_row = None
        for row_num in range(1, sheet.max_row + 1):
            cell_value = self._get_merged_value(sheet, row_num, 2, 5)  # B-E
            if cell_value and str(cell_value).strip().upper() == 'PRODUCTOS SOLICITADOS':
                # Sumamos 3 para empiezar a leer los productos desde la fila 19
                products_start_row = row_num + 3
                break

        if products_start_row:
            current_row = products_start_row
            while current_row <= sheet.max_row:
                # Lee los nombre del producto (columnas B-C)
                name_value = self._get_merged_value(
                    sheet, current_row, 2, 3)  # B-C

                # Si encontramos "Observaciones", terminamos
                if name_value and 'observacion' in str(name_value).lower():
                    # Leer observaciones
                    obs_value = self._get_merged_value(
                        sheet, current_row, 4, 9)  # D-I
                    if not obs_value:
                        obs_value = self._get_merged_value(
                            sheet, current_row + 1, 4, 9)
                    data['observations'] = obs_value
                    break

                # Si la fila está completamente vacía, continuar
                if not name_value:
                    current_row += 1
                    continue

                # Leer cantidad para verificar si es un producto válido
                qty_value = sheet.cell(
                    row=current_row, column=6).value  # F - Cantidad

                if qty_value is not None and qty_value != '':
                    # Es un producto normal
                    try:
                        qty = float(qty_value) if qty_value else 0
                    except (ValueError, TypeError):
                        qty = 0

                    # Leer especificación (columnas D-E)
                    spec_value = self._get_merged_value(
                        sheet, current_row, 4, 5)  # D-E

                    uom = sheet.cell(
                        row=current_row, column=7).value  # G - UdM
                    price_unit = self._get_merged_value(
                        sheet, current_row, 8, 10)  # H-J
                    subtotal = self._get_merged_value(
                        sheet, current_row, 11, 12)  # K-L

                    try:
                        price_unit_float = float(
                            price_unit) if price_unit else 0
                    except (ValueError, TypeError):
                        price_unit_float = 0

                    try:
                        subtotal_float = float(subtotal) if subtotal else 0
                    except (ValueError, TypeError):
                        subtotal_float = 0

                    data['products'].append({
                        'row_num': current_row,
                        'name': str(name_value).strip() if name_value else '',
                        'specification': str(spec_value).strip() if spec_value else '',
                        'qty': qty,
                        'uom': str(uom).strip() if uom else '',
                        'price_unit': price_unit_float,
                        'subtotal': subtotal_float,
                        'is_note': False,
                    })

                current_row += 1

        # Busca el total
        for row_num in range(sheet.max_row, products_start_row if products_start_row else 1, -1):
            total_label = sheet.cell(row=row_num, column=11).value  # K
            if total_label and 'total' in str(total_label).lower():
                total_value = sheet.cell(row=row_num, column=12).value  # L
                try:
                    data['total'] = float(total_value) if total_value else 0
                except (ValueError, TypeError):
                    data['total'] = 0
                break

        # Guarda los datos parseados para mostrar en el resumen
        self.parsed_validity_date = str(
            data['validity_date']) if data['validity_date'] else ''
        self.parsed_currency = str(
            data['currency']) if data['currency'] else ''
        self.parsed_trm = str(data['trm']) if data['trm'] else 'N/A'
        self.parsed_delivery_location = str(
            data['delivery_location']) if data['delivery_location'] else ''
        self.parsed_delivery_date = str(
            data['delivery_date']) if data['delivery_date'] else ''
        self.parsed_includes_freight = str(
            data['includes_freight']) if data['includes_freight'] else ''
        self.parsed_freight_price = str(
            data['freight_price']) if data['freight_price'] else 'N/A'
        self.parsed_observations = str(
            data['observations']) if data['observations'] else ''
        self.parsed_products_count = len(
            [p for p in data['products'] if not p.get('is_note')])
        self.parsed_total = data['total']

        return data

    def _get_merged_value(self, sheet, row, start_col, end_col):
        """Obtiene el valor de una celda, considerando celdas merged"""
        for col in range(start_col, end_col + 1):
            value = sheet.cell(row=row, column=col).value
            if value is not None and value != '':
                return value
        return None

    def _validate_quotation_data(self, data):
        """Valida los datos extraídos del Excel"""
        errors = []
        warnings = []

        # Valida la fecha de vigencia
        if not data['validity_date']:
            errors.append(Markup(
                "La fecha de vigencia <b class='text-primary'>es un campo obligatorio</b> y no se encontró en el archivo."))
        else:
            parsed_date = self._parse_date(data['validity_date'])
            if not parsed_date:
                errors.append(
                    Markup("La fecha de vigencia %s <b class='text-primary'>no tiene un formato válido.</b>") % data['validity_date'])

        # Valida la moneda
        if not data['currency']:
            errors.append(Markup(
                "La moneda <b class='text-primary'>es obligatoria</b> y no se encontró en el archivo."))
        elif data['currency'] not in ['COP', 'USD']:
            errors.append(
                Markup("La moneda %s <b class='text-primary'>no es válida.</b> Debe ser COP o USD.") % data['currency'])

        # Valida el TRM si la moneda es USD
        if data['currency'] == 'USD':
            if not data['trm']:
                errors.append(Markup(
                    "La TRM <b class='text-primary'>es obligatoria</b> cuando la moneda es USD."))
            else:
                try:
                    trm_float = float(data['trm'])
                    if trm_float <= 0:
                        errors.append(Markup(
                            "La TRM <b class='text-primary'>debe ser mayor a 0.</b>"))
                except (ValueError, TypeError):
                    errors.append(
                        Markup("La TRM %s no es válida.") % data['trm'])

        # Valida el lugar de entrega
        if not data['delivery_location']:
            errors.append(Markup(
                "El lugar de entrega <b class='text-primary'>es obligatorio</b> y no se encontró en el archivo."))
        elif data['delivery_location'] not in ['Campo', 'Bodega']:
            warnings.append(
                Markup("El lugar de entrega %s <b class='text-primary'>no es valido (se esperaba 'Campo' o 'Bodega').</b>") % data['delivery_location'])

        # Valida la fecha de entrega
        if not data['delivery_date']:
            errors.append(Markup(
                "La <b class='text-primary'>fecha de entrega</b> es obligatoria y no se encontró en el archivo."))
        else:
            parsed_date = self._parse_date(data['delivery_date'])
            if not parsed_date:
                errors.append(Markup(
                    "La fecha de entrega %s <b class='text-primary'>no tiene un formato válido.</b>") % data['delivery_date'])

        # Valida el flete
        if data['includes_freight'] == 'Sí':
            if not data['freight_price']:
                errors.append(Markup(
                    "Si indica que la cotización incluye flete, <b class='text-primary'>es obligatorio especificar el valor del mismo.</b>"))
            else:
                try:
                    freight_float = float(data['freight_price'])
                    if freight_float <= 0:
                        errors.append(
                            Markup("El precio del flete <b class='text-primary'>debe ser mayor a 0.</b>"))
                except (ValueError, TypeError):
                    errors.append(
                        Markup(
                            "El precio del flete %s <b class='text-primary'>no es válido.</b>") % data['freight_price'])

        # Busca los productos
        products = [p for p in data['products'] if not p.get('is_note')]
        if not products:
            errors.append(Markup(
                "No se encontraron productos en el archivo"
            ))
        else:
            # Busca los productos esperados de la solicitud
            parent = self._get_parent_record()
            expected_products = parent.product_line_ids.filtered(
                lambda l: l.display_type != 'line_note'
            )
            expected_names = {p.name.strip().lower()
                              for p in expected_products}

            for product in products:
                product_name = product['name'].strip().lower()

                # Valida que el producto exista en la solicitud
                if product_name not in expected_names:
                    warnings.append(
                        Markup(
                            "El producto <b class='text-primary'>%s</b> no coincide con los productos de la solicitud.") % product_name.title())

                # Valida el precio unitario
                if product['price_unit'] <= 0:
                    errors.append(
                        Markup(
                            "El producto <b class='text-primary'>%s</b> no tiene un precio unitario válido.") % product_name.title())

                # Valida la especificación
                if not product.get('specification'):
                    errors.append(
                        Markup(
                            "El producto <b class='text-primary'>%s</b> no tiene una especificación.") % product_name.title())

            # Valida que no haya productos duplicados con la misma especificación
            seen_pairs = set()
            for product in products:
                name = product['name'].strip().lower()
                spec = product.get('specification', '').strip().lower()
                pair = (name, spec)
                if pair in seen_pairs:
                    errors.append(
                        Markup(
                            "El producto <b class='text-primary'>%s</b> está duplicado con la misma especificación. "
                            "Los productos duplicados <b class='text-primary'>deben tener especificaciones diferentes.</b>") % product['name'].title())
                seen_pairs.add(pair)

        # Valida el total
        if data['total'] <= 0:
            errors.append(
                Markup(
                    "El total de la cotización <b class='text-primary'>no es válido.</b>"
                ))

        return errors, warnings

    def _parse_date(self, date_value):
        """Valida la fecha en varios formatos"""
        if isinstance(date_value, datetime):
            return date_value.date()

        if not date_value:
            return None

        date_str = str(date_value).strip()

        # Formatos comunes
        formats = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d.%m.%Y',
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return None

    def action_process_file(self):
        """Crea la cotización a partir de los datos del Excel"""
        self.ensure_one()
        if not self.env.user.has_group('purchase_management.group_manager_purchase_management'):
            is_member = self.env['purchase.member'].search_count([
                ('employee_id.user_id', '=', self.env.uid),
            ]) > 0
            if not is_member:
                raise UserError(_('Solo los miembros del área de Compras pueden registrar cotizaciones.'))
        try:
            data = self._analyze_quotation_excel()
            errors, warnings = self._validate_quotation_data(data)

            if errors:
                raise UserError(
                    _('El archivo contiene errores. Por favor corrija y vuelva a intentar.'))

            parent = self._get_parent_record()
            fk_field = self._get_fk_field()
            state = parent.state
            if state not in ['sent', 'in_shopping']:
                raise UserError(
                    _('La solicitud debe estar en estado Enviada para continuar.'))

            # Verifica que el proveedor no tenga ya una cotización (con normalización de nombre)
            duplicate = self._find_duplicate_supplier(parent, self.supplier_name.strip())
            if duplicate:
                raise UserError(
                    _('Ya existe una cotización del proveedor "%s" para esta solicitud.') % duplicate.supplier_name)

            # Crear la cotización
            quotation = self._create_quotation(data)
            parent.sudo().message_post(
                body=Markup(
                    f"<span>Se ha importado una <span style='color: #017e84;'>cotización</span> "
                    f"del proveedor <span style='color: #017e84;'>{self.supplier_name}</span> "
                    f"mediante un archivo Excel.</span>"
                )
            )

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Importación exitosa'),
                    'message': _('Se ha creado la cotización del proveedor %s correctamente.') % self.supplier_name,
                    'type': 'success',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'}
                }
            }
        except Exception as e:
            raise UserError(_('Error al procesar el archivo: %s') % str(e))

    def _create_quotation(self, data):
        """Crea el registro de cotización con sus líneas"""
        self.ensure_one()

        # Busca la moneda
        currency = self.env['res.currency'].search([
            ('name', '=', data['currency'])
        ], limit=1)
        if not currency:
            raise UserError(
                _('La moneda %s no se encuentra en la base de datos. Por favor comuníquese con el administrador del sistema.') % data['currency'])

        # Valida las fechas
        validity_date = self._parse_date(data['validity_date'])
        delivery_date = self._parse_date(data['delivery_date'])

        # Valida el tipo de solicitud
        parent = self._get_parent_record()
        type_id = getattr(parent, 'type_id', False)
        type_string = 'product'
        if type_id and type_id.name == 'Servicios':
            type_string = 'service'
        elif type_id and type_id.name == 'Mixtos':
            type_string = 'mix'

        # TRM
        trm_value = 0
        if data['currency'] == 'USD' and data['trm']:
            try:
                trm_value = float(data['trm'])
            except (ValueError, TypeError):
                trm_value = 0

        # Crea las líneas de productos
        product_lines = []
        products_from_excel = [
            p for p in data['products'] if not p.get('is_note')]

        # Filtra los productos de la solicitud
        request_lines = parent.product_line_ids.filtered(
            lambda l: l.display_type != 'line_note'
        )

        # Rastrear IDs ya usados para que productos con igual nombre se mapeen
        # a líneas distintas de la solicitud (en orden)
        used_request_line_ids = set()

        for excel_product in products_from_excel:
            excel_name = excel_product['name'].strip().lower()

            # Buscar el producto correspondiente en la solicitud (primer no usado)
            matching_line = next(
                (l for l in request_lines
                 if l.name.strip().lower() == excel_name
                 and l.id not in used_request_line_ids),
                None
            )

            if matching_line:
                used_request_line_ids.add(matching_line.id)
                product_lines.append((0, 0, {
                    'request_line_id': matching_line.id,
                    'currency_id': currency.id,
                    'name': matching_line.name,
                    'qty': matching_line.qty,  # Usa la cantidad de la solicitud
                    'uom_id': matching_line.uom_id.id if matching_line.uom_id else False,
                    'price_unit': excel_product['price_unit'],
                    'display_type': False,
                }))

                # Crear línea de especificación desde el campo specification
                if excel_product.get('specification'):
                    product_lines.append((0, 0, {
                        'name': excel_product['specification'],
                        'product_name': matching_line.name,
                        'display_type': 'line_note',
                        'currency_id': currency.id,
                    }))
            else:
                # Si no coincide, crear con el producto con datos del Excel
                product_lines.append((0, 0, {
                    'name': excel_product['name'],
                    'qty': excel_product['qty'],
                    'price_unit': excel_product['price_unit'],
                    'currency_id': currency.id,
                    'display_type': False,
                }))

                # Crea la línea de especificación para productos que no coinciden
                if excel_product.get('specification'):
                    product_lines.append((0, 0, {
                        'name': excel_product['specification'],
                        'product_name': excel_product['name'],
                        'display_type': 'line_note',
                        'currency_id': currency.id,
                    }))

        # Crea la cotización
        fk_field = self._get_fk_field()
        quotation = self.env['quotation.quotation'].sudo().create({
            fk_field: parent.id,
            'supplier_name': self.supplier_name.strip(),
            'currency_id': currency.id,
            'payment_term': '30 días',
            'validity_date': validity_date,
            'delivery_date': delivery_date,
            'type': type_string,
            'trm': trm_value,
            # Se debe actualizar manualmente por el usuario una vez creada la cotizacion
            'evidence': 'https://importado-desde-excel.com',
            'product_line_ids': product_lines,
        })

        # Agrega las observaciones si existen
        if data['observations']:
            self.env['purchase.management.quotation.observation'].sudo().create({
                'quotation_id': quotation.id,
                'name': data['observations'],
            })

        if data['includes_freight'] == 'Sí':
            freight = self.env['purchase.type.commercial.term'].search(
                [('name', '=', 'INCLUYE TRANSPORTE')], limit=1)
            if not freight:
                freight = self.env['purchase.type.commercial.term'].sudo().create({
                    'name': 'INCLUYE TRANSPORTE',
                })
            self.env['purchase.commercial.term'].sudo().create({
                'quotation_id': quotation.id,
                'type_id': freight.id,
            })

        return quotation

    def _generate_error_html(self, errors):
        """Genera el HTML para mostrar los errores"""
        error_rows = ''
        for error in errors:
            error_rows += f'''
            <div class="d-flex align-items-baseline bg-white p-2 mb-2 rounded border">
                <div class="text-primary mr-2 me-2 text-center" style="min-width: 20px;">
                    <i class="fa fa-times-circle" style="font-size: 0.9rem;"></i>
                </div>
                <div class="flex-grow-1">
                    <span class="text-dark" style="font-size: 0.9rem; line-height: 1.5;">{error}</span>
                </div>
            </div>
            '''

        return Markup(f"""
            <div class="bg-white rounded shadow-sm border p-0 overflow-hidden">
                <div class="row no-gutters">
                    <div class="col-md-4 bg-primary text-white d-flex flex-column align-items-center justify-content-center p-4 text-center">
                        <div class="mb-3">
                            <i class="fa fa-exclamation-triangle fa-4x text-white"></i>
                        </div>
                        <h3 class="font-weight-bold m-0 text-white">ERROR</h3>
                        <p class="mb-0 mt-3 text-white" style="opacity: 0.9; font-size: 0.9rem;">
                            Hemos encontrado inconsistencias en los datos. El archivo no puede ser procesado hasta que se realicen las correcciones.
                        </p>
                    </div>
                    <div class="col-md-8 p-0">
                        <div class="px-4 py-3 border-bottom bg-white d-flex justify-content-between align-items-center">
                            <div>
                                <h5 class="text-dark font-weight-bold m-0">Detalle de validación</h5>
                                <small class="text-muted" style="font-size: 0.85rem;">
                                    Revise los registros listados a continuación.
                                </small>
                            </div>
                            <span class="badge badge-primary px-3 py-2" style="font-size: 0.9rem;">
                                {len(errors)} {'error' if len(errors) == 1 else 'errores'}
                            </span>
                        </div>
                        <div class="bg-light" style="height: 280px; overflow-y: auto; border-bottom: 1px solid #dee2e6; padding: 16px 32px 16px 16px;">
                            {error_rows}
                        </div>
                        <div class="px-4 py-2 bg-white text-right text-end">
                            <small class="text-muted font-italic">
                                <i class="fa fa-info-circle text-primary mr-1 me-1"></i>
                                Verifique el archivo Excel y vuelva a cargarlo.
                            </small>
                        </div>
                    </div>
                </div>
            </div>
        """)

    def _generate_success_html(self, data, warnings):
        """Genera el HTML para mostrar que la importación fue correcta"""
        products_count = len(
            [p for p in data['products'] if not p.get('is_note')])

        # Formatea el total
        total_formatted = "{:,.0f}".format(data['total']).replace(',', '.')

        if not warnings:
            return Markup(f"""
                <div class="bg-white rounded shadow-sm border p-0 overflow-hidden">
                    <div class="row no-gutters align-items-stretch">
                        <div class="col-md-4 bg-primary text-white d-flex flex-column align-items-center justify-content-center p-4 text-center">
                            <div class="mb-3">
                                <i class="fa fa-check-circle fa-4x text-white"></i>
                            </div>
                            <h4 class="font-weight-bold text-white">¡Todo Listo!</h4>
                            <p class="small mb-0" style="opacity: 0.9;">
                                La información es correcta y está lista para ser importada.
                            </p>
                        </div>
                        <div class="col-md-8 p-4">
                            <h5 class="text-dark text-uppercase font-weight-bold mb-4"
                                style="font-size: 12px; letter-spacing: 1px;">
                                Resumen de Importación
                            </h5>
                            <div class="row text-center mb-4">
                                <div class="col-4 border-right border-end">
                                    <h2 class="text-primary font-weight-bold m-0">{products_count}</h2>
                                    <small class="text-muted">{'Producto' if products_count == 1 else 'Productos'}</small>
                                </div>
                                <div class="col-4 border-right border-end">
                                    <h4 class="text-primary font-weight-bold m-0">{data['currency']}</h4>
                                    <small class="text-muted">Moneda</small>
                                </div>
                                <div class="col-4">
                                    <h4 class="text-primary font-weight-bold m-0">${total_formatted}</h4>
                                    <small class="text-muted">Total</small>
                                </div>
                            </div>
                            <div class="d-flex align-items-center p-3 rounded" style="background-color: #f8f9fa;">
                                <i class="fa fa-arrow-right text-primary mr-3 me-3 fa-lg"></i>
                                <div>
                                    <span class="d-block text-dark font-weight-bold">Siguiente paso:</span>
                                    <span class="small text-muted">
                                        Haga clic en el botón "Guardar" para crear la cotización.
                                    </span>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            """)

        # Con warnings
        warning_rows = ''
        for w in warnings:
            warning_rows += f"""
                <div class="d-flex align-items-baseline bg-white p-2 mb-2 rounded border w-100">
                    <div class="text-warning mr-2 me-2 text-center" style="min-width: 20px;">
                        <i class="fa fa-exclamation-triangle" style="font-size: 0.9rem;"></i>
                    </div>
                    <div class="flex-grow-1" style="min-width: 0;">
                        <span class="text-dark d-block text-break" style="font-size: 0.9rem; line-height: 1.4;">
                            {w}
                        </span>
                    </div>
                </div>
            """

        return Markup(f"""
            <div class="bg-white rounded shadow-sm border p-0 overflow-hidden">
                <div class="row no-gutters">
                    <div class="col-md-4 bg-primary text-white d-flex flex-column align-items-center justify-content-center p-4 text-center">
                        <div class="mb-3">
                            <i class="fa fa-check-circle fa-4x text-white"></i>
                        </div>
                        <h3 class="font-weight-bold m-0 text-white">¡Todo Listo!</h3>
                        <p class="mb-0 mt-3 text-white" style="opacity: 0.9; font-size: 0.9rem;">
                            Los datos son válidos con algunas observaciones menores.
                        </p>
                    </div>
                    <div class="col-md-8 p-0 d-flex flex-column">
                        <div class="p-3 border-bottom bg-white d-flex justify-content-between align-items-center">
                            <div>
                                <h5 class="text-dark font-weight-bold m-0">Observaciones</h5>
                                <small class="text-muted" style="font-size: 0.85rem;">
                                    Diferencias detectadas con la solicitud.
                                </small>
                            </div>
                            <span class="badge badge-warning px-3 py-2" style="font-size: 0.9rem;">
                                {len(warnings)} {'aviso' if len(warnings) == 1 else 'avisos'}
                            </span>
                        </div>
                        <div class="bg-light flex-grow-1" style="max-height: 200px; overflow-y: auto;">
                            <div style="padding: 16px 32px 16px 16px;">
                                {warning_rows}
                            </div>
                        </div>
                        <div class="bg-white text-right text-end border-top mt-auto"
                             style="padding: 12px 32px 12px 16px;">
                            <small class="text-muted font-italic d-inline-block"
                                   style="font-size: 0.85rem;">
                                <i class="fa fa-cube text-primary mr-1 me-1"></i>
                                {products_count} {'Producto' if products_count == 1 else 'Productos'}
                                <span class="mx-2">|</span>
                                <i class="fa fa-money text-primary mr-1 me-1"></i>
                                ${total_formatted} {data['currency']}
                            </small>
                        </div>
                    </div>
                </div>
            </div>
        """)
