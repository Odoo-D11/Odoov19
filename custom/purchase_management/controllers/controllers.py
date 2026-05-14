from odoo import http, fields
from odoo.http import request
from odoo.exceptions import UserError
import logging
from ..utils.utils import get_distinct_colors, _build_excel_file_supplier, _build_bulk_import_request_template
import base64
import traceback
import io
import json
import unicodedata
import re
import urllib.parse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment


_logger = logging.getLogger(__name__)


class PurchaseManagementController(http.Controller):

    @http.route(
        '/purchase_management/attach_technical_spec',
        type='http', auth='user', methods=['POST'], csrf=True,
    )
    def attach_technical_spec(self, rfq_id, **kwargs):
        import os

        MAX_SIZE = 10 * 1024 * 1024  # 10 MB
        ALLOWED_EXT = {'.pdf', '.doc', '.docx'}
        MIMETYPE_MAP = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        }

        try:
            rfq_id_int = int(rfq_id)
        except (TypeError, ValueError):
            return request.make_json_response({'error': 'ID de solicitud inválido.'}, status=400)

        rfq = request.env['request.quotation'].browse(rfq_id_int)
        if not rfq.exists():
            return request.make_json_response({'error': 'Solicitud no encontrada.'}, status=404)

        uploaded_files = request.httprequest.files.getlist('ufile')
        if not uploaded_files:
            return request.make_json_response({'error': 'No se recibieron archivos.'}, status=400)

        created_ids, errors = [], []
        for ufile in uploaded_files:
            filename = ufile.filename or ''
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ALLOWED_EXT:
                errors.append(
                    f'{filename}: extensión no permitida (solo .pdf, .doc, .docx).')
                continue
            raw_bytes = ufile.read()
            # Compresión de PDF con pikepdf (fallback silencioso si no está disponible)
            if ext == '.pdf':
                try:
                    import pikepdf
                    with pikepdf.open(io.BytesIO(raw_bytes)) as pdf:
                        out_buf = io.BytesIO()
                        pdf.save(out_buf, compress_streams=True,
                                 object_stream_mode=pikepdf.ObjectStreamMode.generate)
                        raw_bytes = out_buf.getvalue()
                except Exception:
                    pass
            if len(raw_bytes) > MAX_SIZE:
                errors.append(f'{filename}: el archivo supera los 10 MB.')
                continue
            att = request.env['ir.attachment'].sudo().create({
                'name': filename,
                'raw': raw_bytes,
                'mimetype': MIMETYPE_MAP.get(ext, 'application/octet-stream'),
                'res_model': 'request.quotation',
                'res_id': rfq_id_int,
            })
            rfq.sudo().write({'technical_spec_attachment_ids': [(4, att.id)]})
            created_ids.append({'id': att.id, 'name': filename})

        if errors and not created_ids:
            return request.make_json_response({'error': '; '.join(errors)}, status=422)

        return request.make_json_response({
            'success': True,
            'attachments': created_ids,
            'warnings': errors,
        })

    @http.route('/purchase_management/get_approval_rights', type='jsonrpc', auth='user')
    def get_approval_rights(self, request_quotation_id=None, **kw):
        try:
            record = request.env['request.quotation'].browse(
                int(request_quotation_id))
            if not record.exists():
                return {'has_rights': False}
            user = request.env.user
            has_rights = (user.has_group('purchase_management.group_purchase_committee_approval')
                          and record.state == 'pending_advisory_committee_approval')
            return {'has_rights': has_rights}
        except Exception as e:
            _logger.error(f"Error checking approval rights: {str(e)}")
            return {'has_rights': False}

    @http.route('/purchase_management/get_comparative_chart_data', type='jsonrpc', auth='user')
    def get_comparative_chart_data(self, request_quotation_id=None, **kw):
        try:
            rfq = request.env['request.quotation'].browse(
                int(request_quotation_id))
            if not rfq.exists():
                return {'error': 'Solicitud no encontrada'}

            quotations = rfq.quotation_line_ids
            requested_products = rfq.product_line_ids
            today = fields.Date.context_today(request.env.user)

            distinct_rgb_list = get_distinct_colors(len(quotations))

            # Dashboard comparativo
            product_map = {}

            # Filtrar solo productos (excluir notas)
            products_only = requested_products.filtered(
                lambda p: not p.display_type)

            for req_prod in products_only:
                product_map[req_prod.id] = {
                    'name': req_prod.name,
                    'requested_qty': int(req_prod.qty),
                    'uom_name': req_prod.uom_id.name if req_prod.uom_id else 'Unidad',
                    'supplier_quotes': {}
                }

            supplier_headers_data = []
            total_price_data = []
            total_price_bg = []
            total_price_border = []
            total_price_complete = []
            total_price_mismatch = []

            delivery_labels = []
            validity_data = []
            validity_bg = []
            validity_border = []
            validity_dates = []

            supplier_list_for_filter = []
            suppliers_info = []
            all_observations_flat = []
            supplier_datasets_info = []

            complete_quote_totals = []

            commercial_info_list = []

            # Determinar proveedores adjudicados (solo cuando la solicitud está aprobada)
            request_state = rfq.state
            awarded_quotation_ids = set()
            if request_state in ('approved', 'order_created'):
                for q in quotations:
                    recent_obs = q.purchase_observation_recommendation_ids.sorted(
                        'create_date', reverse=True)
                    if recent_obs:
                        last_team = recent_obs[0].team
                        team_obs = recent_obs.filtered(
                            lambda o: o.team == last_team)
                        if any(o.state in ('approved', 'recommended') for o in team_obs):
                            awarded_quotation_ids.add(q.id)

            for i, q in enumerate(quotations):
                if q.color and ',' in q.color:
                    rgb_values = q.color
                else:
                    new_color = distinct_rgb_list[i]
                    q.sudo().write({'color': new_color})
                    rgb_values = new_color

                bg_color = f'rgba({rgb_values}, 0.2)'
                border_color = f'rgba({rgb_values}, 1)'

                trm = q.trm if (
                    q.currency_id and q.currency_id.name != 'COP' and q.trm > 0) else 1
                total_cop = int(q.total * trm)
                validity_days = (q.validity_date -
                                 today).days if q.validity_date else None
                validity_str_card = q.validity_date.strftime(
                    '%d/%m/%Y') if q.validity_date else 'N/A'
                delivery_str_table = q.delivery_date.strftime(
                    '%d/%m/%Y') if q.delivery_date else 'No especificada'

                is_complete_quote = True

                has_qty_mismatch = False

                supplier_name = q.supplier_name or ''
                name_parts = supplier_name.split()
                initials = (name_parts[0][0] if name_parts else '') + \
                    (name_parts[1][0] if len(name_parts) > 1 else '')
                is_awarded = q.id in awarded_quotation_ids

                supplier_headers_data.append(
                    {'name': supplier_name, 'is_complete': is_complete_quote, 'is_awarded': is_awarded})

                if is_complete_quote:
                    complete_quote_totals.append(total_cop)

                if validity_days is not None:
                    delivery_labels.append(supplier_name)
                    validity_data.append(validity_days)
                    validity_bg.append(bg_color)
                    validity_border.append(border_color)
                    validity_dates.append(validity_str_card)

                total_price_data.append(total_cop)
                total_price_bg.append(bg_color)
                total_price_border.append(border_color)
                total_price_complete.append(is_complete_quote)
                total_price_mismatch.append(has_qty_mismatch)

                supplier_list_for_filter.append({
                    'id': q.id,
                    'name': supplier_name,
                    'color': border_color
                })

                suppliers_info.append({
                    'id': q.id,
                    'quotation_id': q.id,
                    'name': supplier_name,
                    'initials': initials.upper(),
                    'email': 'No registrado',
                    'evidence': q.evidence or False,
                    'color': bg_color,
                    'border_color': border_color,
                    'is_awarded': is_awarded,
                })

                for obs in q.purchase_observation_recommendation_ids:
                    if obs.team == 'applicant':
                        team_display = 'Proyectos'
                        person_name = rfq.responsible_id.name or False
                        person_badge = 'Líder de Proyecto'
                    elif obs.team == 'final_recommendation':
                        if rfq.category == 'admin':
                            team_display = 'Responsable'
                            person_name = rfq.responsible_id.name or False
                            person_badge = 'Responsable'
                        else:
                            team_display = 'Compras'
                            person_name = rfq.responsible_purchase_id.employee_id.name or False
                            person_badge = 'Compras'
                    else:
                        team_display = 'Comité de Compras'
                        person_name = False
                        person_badge = False
                    all_observations_flat.append({
                        'id': obs.id,
                        'supplier_id': q.id,
                        'supplier_name': supplier_name,
                        'supplier_initials': initials.upper(),
                        'color': bg_color,
                        'border_color': border_color,
                        # Clave original del equipo (applicant, final_recommendation, etc.)
                        'team_key': obs.team,
                        'team': team_display,
                        'person_name': person_name,
                        'person_badge': person_badge,
                        'state': obs.state if obs.state else None,
                        'text': obs.observation,
                        'date': obs.create_date.strftime('%d/%m/%Y') if obs.create_date else '',
                    })

                supplier_datasets_info.append({
                    'id': q.id,
                    'name': supplier_name,
                    'bg': bg_color,
                    'border': border_color
                })

                # Mapear productos de la cotización con líneas de solicitud
                product_lines = q.product_line_ids.filtered(
                    lambda l: not l.display_type)

                for quote_line in product_lines:

                    req_line_id = None

                    # Intentar mapear por request_line_id primero
                    if quote_line.request_line_id and quote_line.request_line_id.id in product_map:
                        req_line_id = quote_line.request_line_id.id

                    # Si no tiene request_line_id, buscar por nombre (fallback para cotizaciones antiguas)
                    elif not quote_line.request_line_id:
                        # Buscar por nombre: preferir el que aún no tenga cotización de este proveedor
                        for prod_id, prod_data in product_map.items():
                            if prod_data['name'] and quote_line.name and prod_data['name'].strip().lower() == quote_line.name.strip().lower():
                                if q.id not in prod_data['supplier_quotes']:
                                    req_line_id = prod_id
                                    break
                                elif req_line_id is None:
                                    req_line_id = prod_id  # Guardar como fallback si todos ya tienen datos

                        if not req_line_id:
                            _logger.warning(
                                f"NO se pudo mapear por nombre: '{quote_line.name}'")

                    # Si se encontró un mapeo (por ID o por nombre), agregar los datos
                    if req_line_id:
                        price_cop = quote_line.subtotal * trm
                        unit_price_cop = float(quote_line.price_unit) * trm
                        requested_qty = product_map[req_line_id]['requested_qty']
                        qty_mismatch = (int(quote_line.qty) != requested_qty)

                        # Buscar la nota/especificación asociada a este producto
                        note_text = ''
                        all_lines = q.product_line_ids.sorted(
                            key=lambda l: l.sequence)
                        for i, line in enumerate(all_lines):
                            if line.id == quote_line.id:
                                # Buscar la siguiente línea que sea una nota con el mismo product_name
                                if i + 1 < len(all_lines):
                                    next_line = all_lines[i + 1]
                                    if next_line.display_type == 'line_note' and next_line.product_name == quote_line.name:
                                        note_text = next_line.name or ''
                                break

                        product_map[req_line_id]['supplier_quotes'][q.id] = {
                            'quoted_qty': int(quote_line.qty),
                            'unit_price': float(quote_line.price_unit),
                            'unit_price_cop': unit_price_cop,
                            'total_price': float(quote_line.subtotal),
                            'normalized_price': price_cop,
                            'qty_mismatch': qty_mismatch,
                            'note': note_text
                        }
                    else:
                        if quote_line.request_line_id:
                            _logger.warning(
                                f"request_line_id {quote_line.request_line_id.id} no está en product_map (claves: {list(product_map.keys())})")

                delivery_days_diff = (
                    q.delivery_date - today).days if q.delivery_date else None
                payment_term_name = q.payment_term or 'No especificado'
                currency_name = q.currency_id.name if q.currency_id else 'COP'

                comm_project = q.request_quotation_id.project_id.name if q.request_quotation_id and q.request_quotation_id.project_id else 'N/A'
                comm_reference = q.request_quotation_id.reference if q.request_quotation_id else 'N/A'

                commercial_info_list.append({
                    'id': q.id,
                    'name': supplier_name,
                    'initials': initials.upper(),
                    'color': bg_color,
                    'border_color': border_color,
                    'delivery_date': delivery_str_table,
                    'delivery_days': delivery_days_diff,
                    'payment_term': payment_term_name,
                    'currency': currency_name,
                    'trm': q.trm,
                    'email': '',
                    'project': comm_project,
                    'reference': comm_reference,
                    'is_awarded': is_awarded,
                })

            table_rows = []
            best_combination_total = 0

            supplier_datasets = []
            for info in supplier_datasets_info:
                supplier_datasets.append({
                    'label': info['name'],
                    'data': [], 'is_best': [], 'qty_mismatch': [],
                    'backgroundColor': info['bg'], 'borderColor': info['border'],
                    'borderWidth': 2, 'fill': False, 'tension': 0.1
                })

            for prod_id, data in product_map.items():
                prices = [quote['normalized_price'] for quote in data['supplier_quotes'].values(
                ) if quote['normalized_price'] > 0]
                min_price = min(prices) if prices else 0
                best_combination_total += min_price

                row = {
                    'product_id': prod_id,
                    'product_name': data['name'],
                    'requested_qty': data['requested_qty'],
                    'uom_name': data['uom_name'],
                    'supplier_cells': [],
                    'best_price': min_price
                }

                for i, info in enumerate(supplier_datasets_info):
                    quote = data['supplier_quotes'].get(info['id'])

                    cell_data = {'total_price': 0, 'unit_price': 0,
                                 'is_best': False, 'qty_mismatch': False, 'quoted_qty': 0}
                    val_normalized = None
                    is_best = False
                    mismatch = False

                    if quote:
                        is_best = (quote['normalized_price'] ==
                                   min_price and min_price > 0)
                        cell_data.update(quote)
                        cell_data['is_best'] = is_best
                        val_normalized = int(quote['normalized_price'])
                        mismatch = quote['qty_mismatch']

                    row['supplier_cells'].append(cell_data)

                    supplier_datasets[i]['data'].append(val_normalized)
                    supplier_datasets[i]['is_best'].append(is_best)
                    supplier_datasets[i]['qty_mismatch'].append(mismatch)

                table_rows.append(row)

            """FECHA DE ENTREGA MÁS RÁPIDA"""

            fastest_delivery_days = None

            for q in quotations:
                validity_days_check = (
                    q.validity_date - today).days if q.validity_date else -1
                is_valid = validity_days_check >= 0
                if is_valid and q.delivery_date:
                    days_to_delivery = (q.delivery_date - today).days
                    if fastest_delivery_days is None or days_to_delivery < fastest_delivery_days:
                        fastest_delivery_days = days_to_delivery

            fastest_delivery_str = "N/A"
            if fastest_delivery_days is not None:
                suffix = "Día" if fastest_delivery_days == 1 else "Días"
                if fastest_delivery_days == 0:
                    fastest_delivery_str = "Hoy"
                elif fastest_delivery_days < 0:
                    fastest_delivery_str = f"Hace {abs(fastest_delivery_days)} {suffix}"
                else:
                    fastest_delivery_str = f"{fastest_delivery_days} {suffix}"

            min_total_cop = min(
                complete_quote_totals) if complete_quote_totals else 0
            valid_validity_days = [d for d in validity_data if d is not None]
            max_validity = max(
                valid_validity_days) if valid_validity_days else None
            kpi_validity_str = f"{max_validity} Día{'s' if max_validity != 1 else ''}" if max_validity is not None else "N/A"

            kpis = {
                'total_quotations': len(quotations),
                'fastest_delivery_time': fastest_delivery_str,
                'best_validity': kpi_validity_str,
                'best_combination_total': int(best_combination_total)
            }

            all_product_details = []
            for prod_id, data in product_map.items():
                details_map = {
                    'null': {'requested_qty': data['requested_qty'], 'quoted_qty': None, 'uom': data['uom_name']}
                }
                for info in supplier_datasets_info:
                    q_data = data['supplier_quotes'].get(info['id'], {})
                    details_map[str(info['id'])] = {
                        'requested_qty': data['requested_qty'],
                        'quoted_qty': q_data.get('quoted_qty', 0),
                        'uom': data['uom_name']
                    }
                all_product_details.append(details_map)

            product_comparison_data = {
                'labels': [p['name'] for p in product_map.values()],
                'product_ids': list(product_map.keys()),
                'all_product_details': all_product_details,
                'suppliers_ids': [info['id'] for info in supplier_datasets_info],
                'datasets': supplier_datasets
            }

            return {
                'kpis': kpis,
                'tableData': {
                    'supplier_headers': supplier_headers_data,
                    'rows': table_rows,
                    'totals': {
                        'supplier_totals': total_price_data,
                        'best_total_complete': int(min_total_cop),
                        'best_total_combination': int(best_combination_total)
                    },
                },
                'totalPriceChartData': {
                    'labels': [h['name'] for h in supplier_headers_data],
                    'datasets': [{
                        'label': 'Total Cotizado',
                        'data': total_price_data,
                        'backgroundColor': total_price_bg,
                        'borderColor': total_price_border,
                        'borderWidth': 1,
                        'is_complete': total_price_complete,
                        'has_qty_mismatch': total_price_mismatch
                    }]
                },
                'validityChartData': {
                    'labels': delivery_labels,
                    'datasets': [{
                        'label': 'Días Vigencia',
                        'data': validity_data,
                        'backgroundColor': validity_bg,
                        'borderColor': validity_border,
                        'borderWidth': 1
                    }],
                    'validity_dates': validity_dates,
                    'has_data': bool(validity_data)
                },
                'productComparisonData': product_comparison_data,
                'productList': [{'id': pid, 'name': d['name']} for pid, d in product_map.items()],
                'supplierList': supplier_list_for_filter,
                'suppliersInfo': suppliers_info,
                'allObservations': all_observations_flat,
                'commercialTerms': commercial_info_list,
                'rfq_state': request_state,
            }

        except Exception as e:
            tb_str = traceback.format_exc()
            _logger.error(
                f"Error al generar datos del dashboard: {str(e)}\n{tb_str}")
            return {'error': f"Error en el controlador: {str(e)}"}

    @http.route('/purchase_management/get_products_for_supplier', type='jsonrpc', auth='user')
    def get_products_for_supplier(self, request_quotation_id=None, **kw):
        """Obtiene los productos de una solicitud para enviar a proveedores"""
        try:
            record = request.env['request.quotation'].browse(
                int(request_quotation_id))
            type_name = record.type_id.name if record.type_id else 'Productos'

            if not record.exists():
                return {'error': 'Solicitud no encontrada'}

            products_only = record.product_line_ids.filtered(
                lambda p: not p.display_type).sorted('sequence')

            products = []
            for product in products_only:
                # Busca la especificación asociada al producto
                spec = record.product_line_ids.filtered(
                    lambda p: p.display_type == 'line_note' and p.product_name == product.name
                )
                spec_text = spec[0].name if spec else ''

                products.append({
                    'id': product.id,
                    'name': product.name,
                    'qty': int(product.qty),
                    'uom': product.uom_id.name if product.uom_id else 'Unidad',
                    'specification': spec_text,
                    'sent_to_emails': product.sent_to_emails or '',
                })

            # Datos del usuario actual
            current_user = request.env.user

            return {
                'request_quotation_id': record.id,
                'reference': record.reference,
                'subject': record.subject,
                'products': products,
                'total_products': len(products),
                'type_name': type_name,
                'user_name': current_user.name,
                'user_email': current_user.email or current_user.login,
            }

        except Exception as e:
            _logger.error(
                f"Error al obtener productos para proveedor: {str(e)}")
            return {'error': str(e)}

    @http.route('/purchase_management/create_excel_for_selected_products', type='jsonrpc', auth='user')
    def create_excel_for_selected_products(self, request_quotation_id=None, product_ids=None, **kw):
        """Genera el Excel con los productos seleccionados y retorna el attachment_id"""
        try:
            record = request.env['request.quotation'].browse(
                int(request_quotation_id))

            if not record.exists():
                return {'error': 'Solicitud no encontrada'}

            if not product_ids:
                return {'error': 'Debe seleccionar al menos un producto'}

            # Generar el Excel con los productos seleccionados
            file_content = _build_excel_file_supplier(record, product_ids)
            attachment = request.env['ir.attachment'].sudo().create({
                'name': f"{record.reference}.xlsx",
                'datas': base64.b64encode(file_content),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'res_model': 'purchase.send.mail.wizard',
                'res_id': 0,
            })

            return {
                'success': True,
                'attachment_id': attachment.id,
                'attachment_name': attachment.name,
            }

        except Exception as e:
            _logger.error(
                f"Error al crear Excel para productos seleccionados: {str(e)}\n{traceback.format_exc()}")
            return {'error': str(e)}

    @http.route('/purchase_management/get_order_mapping_data', type='jsonrpc', auth='user')
    def get_order_mapping_data(self, quotation_id=None, **kw):
        """Retorna las líneas de producto de una solicitud aprobada con sugerencias
        de warehouse.product.variant basadas en coincidencia de nombre."""
        try:
            record = request.env['request.quotation'].browse(int(quotation_id))
            if not record.exists():
                return {'error': 'Solicitud no encontrada'}

            product_lines = record.product_line_ids.filtered(
                lambda l: not l.display_type
            ).sorted('sequence')

            WVariant = request.env['warehouse.product.variant'].sudo()
            WAlias = request.env['warehouse.variant.alias'].sudo()

            # Índice de especificaciones: product_name -> texto de la nota
            note_lines = record.product_line_ids.filtered(
                lambda l: l.display_type == 'line_note'
            )
            spec_by_pname = {}
            for nl in note_lines:
                key = nl.product_name or ''
                if key and key not in spec_by_pname:
                    spec_by_pname[key] = (nl.name or '').strip()

            # Cotizaciones adjudicadas para completar automaticamente el precio estimado en el modal.
            # Opción 1: campo awarded=True en el modelo (Depende segun lo marcado por el comite de compras).
            # Opción 2 lógica de observaciones igual al cuadro comparativo.
            awarded_quotations = [
                q for q in record.quotation_line_ids if q.awarded]

            if not awarded_quotations:
                # Fallback: buscar cotizaciones que el comité de compras aprobó o recomendó.
                # Se filtra explícitamente por team='purchase_committee' porque cuando la
                # solicitud llega a estado 'approved' la última instancia que votó es el comité.
                for q in record.quotation_line_ids:
                    committee_obs = q.purchase_observation_recommendation_ids.filtered(
                        lambda o: o.team == 'purchase_committee'
                    )
                    if any(o.state in ('approved', 'recommended') for o in committee_obs):
                        awarded_quotations.append(q)

            lines = []
            for line in product_lines:
                # Buscar coincidencias en variante y producto padre
                raw_name = (line.name or '').strip()
                tokens = [t for t in raw_name.split() if len(t) > 2]

                if tokens:
                    from odoo.fields import Domain as expr
                    domain = expr.OR(
                        [[('name', 'ilike', t)] for t in tokens] +
                        [[('product_id.name', 'ilike', t)] for t in tokens]
                    )
                else:
                    domain = ['|', ('name', 'ilike', raw_name),
                              ('product_id.name', 'ilike', raw_name)] if raw_name else []

                suggestions = []
                if domain:
                    line_words = set(t.lower() for t in tokens)
                    for v in WVariant.search(domain, limit=15):
                        v_product = (v.product_id.name or '').lower(
                        ) if v.product_id else ''
                        v_name = (v.name or '').lower()
                        v_words = set(w for w in (
                            v_product + ' ' + v_name).split() if len(w) > 2)
                        if line_words and v_words:
                            common = len(line_words & v_words)
                            score = common / max(len(line_words), len(v_words))
                        else:
                            score = 0
                        if score >= 0.4:
                            suggestions.append({
                                'id': v.id,
                                'name': v.name,
                                'productName': v.product_id.name if v.product_id else '',
                                'category': v.product_id.category_id.name if v.product_id and v.product_id.category_id else '',
                            })
                            if len(suggestions) >= 5:
                                break

                # Completar automaticamente con el alias exacto (si existe)
                preselected = None
                alias_v = self._find_variant_by_alias(WAlias, raw_name)
                if alias_v:
                    preselected = {
                        'id': alias_v.id,
                        'name': alias_v.name,
                        'productName': alias_v.product_id.name if alias_v.product_id else '',
                        'category': alias_v.product_id.category_id.name
                        if alias_v.product_id and alias_v.product_id.category_id else '',
                    }
                # 1. Sin no existe el alias buscar la mejor coincidencia con el nombre completo (score >= 0.5)
                elif suggestions:
                    line_words_set = set(t.lower() for t in tokens)
                    best_score = 0.0
                    for s in suggestions:
                        s_words = set(
                            w for w in (
                                (s['productName'] or '') +
                                ' ' + (s['name'] or '')
                            ).lower().split()
                            if len(w) > 1
                        )
                        if line_words_set and s_words:
                            common = len(line_words_set & s_words)
                            score = common / \
                                max(len(line_words_set), len(s_words))
                            if score > best_score:
                                best_score = score
                                if score >= 0.5:
                                    preselected = s

                # Para cada cotización adjudicada, buscar la línea que corresponde a esta
                # línea de solicitud y construir la lista de proveedores adjudicados.
                Partner = request.env['res.partner'].sudo()
                awarded_suppliers = []
                for aq in awarded_quotations:
                    is_foreign = (
                        aq.currency_id and aq.currency_id.name != 'COP' and aq.trm > 0)
                    aq_trm = aq.trm if is_foreign else 1
                    for qline in aq.product_line_ids.filtered(lambda l: not l.display_type):
                        matched = False
                        if qline.request_line_id and qline.request_line_id.id == line.id:
                            matched = True
                        elif not qline.request_line_id and qline.name and raw_name and \
                                qline.name.strip().lower() == raw_name.lower():
                            matched = True
                        if matched:
                            unit_price_raw = float(qline.price_unit)
                            unit_price_cop = int(unit_price_raw * aq_trm)
                            supplier_name = aq.supplier_name or ''
                            existing_partner = Partner.search(
                                [('name', '=ilike', supplier_name.strip())], limit=1
                            ) if supplier_name.strip() else Partner
                            awarded_suppliers.append({
                                'quotation_id': aq.id,
                                'qline_id': qline.id,
                                'partner_name': supplier_name,
                                'unit_price': unit_price_raw,
                                'unit_price_cop': unit_price_cop,
                                'currency_name': aq.currency_id.name if aq.currency_id else 'COP',
                                'currency_id': aq.currency_id.id if aq.currency_id else False,
                                'trm': float(aq_trm),
                                'assigned_qty': qline.assigned_qty or 0,
                                'partner_found': bool(existing_partner),
                                'partner_id': existing_partner.id if existing_partner else False,
                                'payment_term_name': aq.payment_term or '',
                            })
                            break

                # awarded_unit_price_cop: precio del primer proveedor adjudicado (compatibilidad)
                awarded_unit_price_cop = awarded_suppliers[0]['unit_price_cop'] if awarded_suppliers else 0

                lines.append({
                    'id': line.id,
                    'name': raw_name,
                    'qty': int(line.qty),
                    'uom': line.uom_id.name if line.uom_id else 'Unidad',
                    'warehouse_variant_id': line.warehouse_variant_id.id or False,
                    'suggestions': suggestions,
                    'preselected_variant_id': preselected['id'] if preselected else False,
                    'preselected_variant_name': preselected['name'] if preselected else '',
                    'preselected_product_name': preselected['productName'] if preselected else '',
                    'preselected_category': preselected['category'] if preselected else '',
                    'specification': spec_by_pname.get(line.product_name or raw_name, ''),
                    'cost_center_id': line.cost_center_id.id or False,
                    'cost_center_name': line.cost_center_id.display_name if line.cost_center_id else '',
                    'awarded_unit_price_cop': awarded_unit_price_cop,
                    'awarded_suppliers': awarded_suppliers,
                })

            # Opciones para el modal de creación de producto
            WCategory = request.env['warehouse.product.category'].sudo()
            WType = request.env['type.product'].sudo()

            categories = [
                {'id': c.id, 'name': c.name}
                for c in WCategory.search([], order='name')
            ]
            product_types = [
                {'id': t.id, 'name': t.name}
                for t in WType.search([], order='name')
            ]

            # Centro de costo por defecto: "Suministro de materiales" del proyecto.
            # Usa sudo() y normalización NFD para evitar problemas de acceso y tildes.
            default_cc = False
            if record.project_id:
                _target = unicodedata.normalize('NFD', 'Suministro de materiales') \
                    .encode('ascii', 'ignore').decode('ascii').lower().strip()
                for cc in record.project_id.sudo().cost_center_ids:
                    if not cc.analytical_account_id:
                        continue
                    _cc_name = unicodedata.normalize('NFD', cc.analytical_account_id.name or '') \
                        .encode('ascii', 'ignore').decode('ascii').lower().strip()
                    if _cc_name == _target:
                        default_cc = cc
                        break

            payment_terms = [
                {'id': t.id, 'name': t.name}
                for t in request.env['accounting.payment.term'].sudo().search([], order='name')
            ]
            currencies = [
                {'id': c.id, 'name': c.name}
                for c in request.env['res.currency'].sudo().search(
                    [('active', '=', True)], order='name'
                )
            ]
            colombia = request.env['res.country'].sudo().search(
                [('code', '=', 'CO')], limit=1)
            colombian_states = [
                {'id': s.id, 'name': s.name}
                for s in request.env['res.country.state'].sudo().search(
                    [('country_id', '=', colombia.id)], order='name'
                )
            ] if colombia else []

            return {
                'reference': record.reference,
                'subject': record.subject,
                'request_type': record.type_id.name or '',
                'project_id': record.project_id.id or False,
                'lines': lines,
                'categories': categories,
                'product_types': product_types,
                'default_cost_center_id': default_cc.id if default_cc else False,
                'default_cost_center_name': default_cc.display_name if default_cc else '',
                'payment_terms': payment_terms,
                'currencies': currencies,
                'colombian_states': colombian_states,
            }

        except Exception as e:
            _logger.error(
                f"Error al obtener datos de mapeo de orden: {str(e)}\n{traceback.format_exc()}")
            return {'error': str(e)}

    @http.route('/purchase_management/create_and_confirm_mapping', type='jsonrpc', auth='user')
    def create_and_confirm_mapping(self, quotation_id=None, mappings=None,
                                   distributions=None, pending_partners=None,
                                   delivery_data=None, **kw):
        """Tiene toda la lógica de creación de OC al modelo request.quotation.
        Buscar request.quotation.action_create_purchase_orders() para la documentacion
        completa de los parámetros mappings y distributions.
        """
        try:
            import re as _re
            record = request.env['request.quotation'].browse(int(quotation_id))
            if not record.exists():
                return {'error': 'Solicitud no encontrada'}

            with request.env.cr.savepoint():
                # Crear los partners pendientes antes de generar las OC
                for pp in (pending_partners or []):
                    supplier_name = (pp.get('supplier_name') or '').strip()
                    if not supplier_name:
                        continue
                    Partner = request.env['res.partner'].sudo()
                    # Verificar NIT duplicado
                    nit = (pp.get('nit') or '').strip()
                    if nit:
                        clean_nit = _re.sub(r'[.\-\s]', '', nit)
                        for existing in Partner.search([('vat', '!=', False)]):
                            if _re.sub(r'[.\-\s]', '', existing.vat or '') == clean_nit:
                                raise UserError(
                                    'El NIT %s ya está registrado para "%s".'
                                    % (nit, existing.name))
                    colombia = request.env['res.country'].sudo().search(
                        [('code', '=', 'CO')], limit=1)
                    nit_type = request.env['identification.type'].sudo().search(
                        [('name', '=', 'NIT')], limit=1)
                    vals = {
                        'name': supplier_name.title(),
                        'is_supplier': True,
                        'is_company': True,
                        'country_id': colombia.id if colombia else False,
                    }
                    if nit_type:
                        vals['identification_type_id'] = nit_type.id
                    if pp.get('email'):
                        vals['email'] = pp['email'].strip()
                    if pp.get('phone'):
                        vals['phone'] = pp['phone'].strip()
                    if nit:
                        vals['vat'] = nit
                    if pp.get('address'):
                        vals['street'] = pp['address'].strip()
                    city_id = pp.get('city_id')
                    if city_id:
                        city = request.env['res.city'].sudo().browse(
                            int(city_id))
                        if city.exists():
                            vals['city_id'] = city.id
                            if city.state_id:
                                vals['state_id'] = city.state_id.id
                    # Plazo y moneda desde la cotización
                    quot_id = pp.get('quotation_id')
                    if quot_id:
                        quot = request.env['quotation.quotation'].sudo().browse(
                            int(quot_id))
                        if quot.exists():
                            if quot.payment_term:
                                vals['payment_term'] = quot.payment_term
                            if quot.currency_id:
                                vals['currency_id'] = quot.currency_id.id
                    cat_ids = pp.get('category_ids') or []
                    if cat_ids:
                        vals['product_category_ids'] = [
                            (6, 0, [int(c) for c in cat_ids])]
                    Partner.with_company(request.env.company).create(vals)

                created_variants = record.action_create_purchase_orders(
                    mappings or [],
                    distributions or [],
                    delivery_data=delivery_data or {},
                )

            return {'success': True, 'created_variants': created_variants}

        except UserError as e:
            return {'error': str(e)}
        except Exception as e:
            _logger.error(
                'Error en create_and_confirm_mapping: %s\n%s', e, traceback.format_exc())
            return {'error': str(e)}

    @http.route('/purchase_management/search_cities', type='jsonrpc', auth='user')
    def search_cities(self, search_term='', state_id=None, limit=10, **kw):
        try:
            domain = [('active', '=', True)]
            if search_term:
                domain += [('name', 'ilike', search_term)]
            if state_id:
                domain += [('state_id', '=', int(state_id))]
            cities = request.env['res.city'].sudo().search(
                domain, limit=limit, order='name')
            return {'cities': [{'id': c.id, 'name': c.name, 'state_id': c.state_id.id, 'state_name': c.state_id.name} for c in cities]}
        except Exception as e:
            return {'error': str(e)}

    @http.route('/purchase_management/check_nit', type='jsonrpc', auth='user')
    def check_nit(self, nit='', **kw):
        try:
            import re
            clean = re.sub(r'[.\-\s]', '', (nit or '').strip())
            if not clean:
                return {'exists': False}
            partners = request.env['res.partner'].sudo().search(
                [('vat', '!=', False)])
            for p in partners:
                if re.sub(r'[.\-\s]', '', p.vat or '') == clean:
                    return {'exists': True, 'partner_name': p.name}
            return {'exists': False}
        except Exception as e:
            return {'error': str(e)}

    @http.route('/purchase_management/create_supplier_partner', type='jsonrpc', auth='user')
    def create_supplier_partner(self, supplier_name='', email='', phone='', nit='',
                                city_id=None, quotation_id=None, category_ids=None, **kw):
        try:
            if not (supplier_name or '').strip():
                return {'error': 'El nombre del proveedor es obligatorio.'}
            Partner = request.env['res.partner'].sudo()
            existing = Partner.search(
                [('name', '=ilike', supplier_name.strip())], limit=1
            )
            if existing:
                return {'partner_id': existing.id, 'partner_name': existing.name, 'already_exists': True}
            colombia = request.env['res.country'].sudo().search(
                [('code', '=', 'CO')], limit=1)
            nit_type = request.env['identification.type'].sudo().search(
                [('name', '=', 'NIT')], limit=1)
            vals = {
                'name': supplier_name.strip().title(),
                'is_supplier': True,
                'is_company': True,
                'country_id': colombia.id if colombia else False,
            }
            if nit_type:
                vals['identification_type_id'] = nit_type.id
            if email and email.strip():
                vals['email'] = email.strip()
            if phone and phone.strip():
                vals['phone'] = phone.strip()
            if nit and nit.strip():
                vals['vat'] = nit.strip()
            if city_id:
                city = request.env['res.city'].sudo().browse(int(city_id))
                if city.exists():
                    vals['city_id'] = city.id
                    if city.state_id:
                        vals['state_id'] = city.state_id.id
            # Tomar plazo de pago y moneda de la cotización
            if quotation_id:
                quotation = request.env['quotation.quotation'].sudo().browse(
                    int(quotation_id))
                if quotation.exists():
                    if quotation.payment_term:
                        vals['payment_term'] = quotation.payment_term
                    if quotation.currency_id:
                        vals['currency_id'] = quotation.currency_id.id
            if category_ids:
                vals['product_category_ids'] = [
                    (6, 0, [int(c) for c in category_ids])]
            partner = Partner.with_company(request.env.company).create(vals)
            return {'partner_id': partner.id, 'partner_name': partner.name, 'success': True}
        except Exception as e:
            _logger.error('Error en create_supplier_partner: %s\n%s',
                          e, traceback.format_exc())
            return {'error': str(e)}

    @http.route('/purchase_management/search_warehouse_products', type='jsonrpc', auth='user')
    def search_warehouse_products(self, search_term='', limit=10, product_id=None, **kw):
        """Busca variantes en el catalogo de bodega por nombre o SKU.
        Con term vacío retorna los primeros `limit` resultados ordenados alfabéticamente,
        lo que permite mostrar el dropdown de inmediato al enfocar el campo.

        product_id (int, opcional): filtra las variantes al producto indicado.
        Usado por el modal de creación para detectar variantes duplicadas.
        """
        try:
            WVariant = request.env['warehouse.product.variant'].sudo()
            term = (search_term or '').strip()
            if term:
                domain = ['|', '|', '|',
                          ('name', 'ilike', term),
                          ('sku', 'ilike', term),
                          ('product_id.name', 'ilike', term),
                          ('product_id.sku', 'ilike', term),
                          ]
            else:
                domain = []
            if product_id:
                domain = [('product_id', '=', int(product_id))] + domain
            matches = WVariant.search(domain, limit=limit, order='name asc')

            products = []
            for v in matches:
                products.append({
                    'id': v.id,
                    'name': v.name,
                    'productName': v.product_id.name if v.product_id else '',
                    'category': v.product_id.category_id.name if v.product_id and v.product_id.category_id else '',
                })

            return {'products': products}

        except Exception as e:
            _logger.error(
                f"Error al buscar variantes en bodega: {str(e)}\n{traceback.format_exc()}")
            return {'error': str(e)}

    @http.route('/purchase_management/search_cost_centers', type='jsonrpc', auth='user')
    def search_cost_centers(self, search_term='', limit=12, project_id=None, request_type='', **kw):
        """Busca centros de costo filtrando por proyecto y, si la solicitud es de tipo
        Productos, solo muestra Suministro de Equipos (42) y Suministro de materiales (46)."""
        try:
            CostCenter = request.env['cost.center'].sudo()
            term = (search_term or '').strip()
            domain = []

            # Limitar siempre al proyecto de la solicitud
            if project_id:
                domain.append(('project_id', '=', int(project_id)))

            # Para solicitudes de tipo Productos solo aplican Suministro de Equipos y Suministro de materiales
            if (request_type or '').strip().lower() == 'productos':
                domain.append(('analytical_account_id.name', 'in', [
                              'Suministro de Equipos', 'Suministro de materiales']))

            # Busca por texto en código o nombre del centro de costo o su cuenta analítica
            if term:
                text_domain = ['|', '|',
                               ('code', 'ilike', term),
                               ('analytical_account_id.name', 'ilike', term),
                               ('analytical_account_id.code', 'ilike', term),
                               ]
                domain = domain + text_domain if domain else text_domain

            records = CostCenter.search(domain, limit=limit, order='code asc')
            cost_centers = [{'id': cc.id, 'name': cc.display_name}
                            for cc in records]
            return {'cost_centers': cost_centers}
        except Exception as e:
            _logger.error(
                f"Error al buscar centros de costo: {str(e)}\n{traceback.format_exc()}")
            return {'error': str(e)}

    @http.route('/purchase_management/search_warehouse_product_templates', type='jsonrpc', auth='user')
    def search_warehouse_product_templates(self, search_term='', limit=8, **kw):
        """Busca warehouse.product (plantillas) por nombre o SKU.
        Usado en el modal de creación cuando el usuario elige el modo 'Variante'."""
        try:
            WProduct = request.env['warehouse.product'].sudo()
            term = (search_term or '').strip()
            domain = ['|', ('name', 'ilike', term),
                      ('sku', 'ilike', term)] if term else []
            matches = WProduct.search(domain, limit=limit, order='name asc')
            return {'products': [
                {
                    'id': p.id,
                    'name': p.name,
                    'category': p.category_id.name if p.category_id else '',
                }
                for p in matches
            ]}
        except Exception as e:
            _logger.error(
                f"Error al buscar plantillas de producto: {str(e)}\n{traceback.format_exc()}")
            return {'error': str(e)}

    @http.route('/purchase_management/suggest_description', type='jsonrpc', auth='user')
    def suggest_description(self, product_name='', **kw):
        """Busca una descripción del producto consultando Wikipedia en español (primary)
        y DuckDuckGo Instant Answer (fallback). Actúa como proxy para evitar CORS.
        Devuelve {'description': str, 'source': str} o {'description': '', 'error': str}."""
        import requests as _requests

        product_name = (product_name or '').strip()
        if not product_name:
            return {'description': '', 'error': 'Nombre de producto vacío.'}

        _ALNUM_RE = re.compile(r'[^a-zA-ZáéíóúÁÉÍÓÚñÑüÜ0-9]')
        _HEADERS = {'User-Agent': 'OdooProductDescSuggest/1.0 (internal tool)'}

        def _truncate_to_limit(text, limit=200):
            count = 0
            for i, ch in enumerate(text):
                if not _ALNUM_RE.match(ch):
                    continue
                count += 1
                if count > limit:
                    cut = text.rfind(' ', 0, i)
                    return text[:cut].rstrip() if cut > 0 else text[:i]
            return text

        def _clean_extract(text):
            """Limpia el texto de Wikipedia eliminando ruido enciclopédico:
            prefijos de campo, listas de sinónimos y paréntesis de traducción."""
            # 1. Quitar "En [campo], " → el sujeto queda capitalizable
            text = re.sub(
                r'^En [a-záéíóúüñA-ZÁÉÍÓÚÜÑ][a-záéíóúüñA-ZÁÉÍÓÚÜÑ\s\-]+,\s+',
                '', text,
            )
            # 2. Quitar lista de sinónimos parando justo antes del verbo principal
            #    "X, también llamado A, B, C, es..." → "X, es..."
            text = re.sub(
                r',?\s+(?:también\s+)?(?:llamad[oa]|conocid[oa]|denominad[oa]|abreviad[oa])\s+'
                r'.+?(?=,?\s+(?:es|son|fue|era|puede|sirve|permite|se\s)\b)',
                '', text, flags=re.IGNORECASE,
            )
            # 3. Limpiar coma sobrante ante verbo: ", es..." → " es..."
            text = re.sub(
                r',\s+(?=(?:es|son|fue|era|puede|sirve|permite|se)\s)',
                ' ', text, flags=re.IGNORECASE,
            )
            # 4. Quitar paréntesis de traducción: (del inglés: X), (en inglés: X)
            text = re.sub(
                r'\s*\((?:también[^)]+|(?:del|en)\s+inglés[^)]+)\)',
                '', text, flags=re.IGNORECASE,
            )
            # 5. Conservar solo las primeras 1-2 oraciones
            parts = re.split(r'(?<=[.!?])\s+(?=[A-ZÁÉÍÓÚ«"¿])', text)
            result = parts[0] if len(parts[0]) > 80 else ' '.join(parts[:2])
            # 6. Limpiar espacios dobles y puntuación suelta
            result = re.sub(r'\s+', ' ', result).strip()
            result = re.sub(r'\s+([,.])', r'\1', result)
            result = re.sub(r',\s*\.', '.', result)
            if result and result[-1] not in '.!?':
                result += '.'
            # 7. Capitalizar primera letra (tras haber quitado el prefijo "En X,")
            return result[0].upper() + result[1:] if result else result

        # ── 0. Ollama (Piloto) ────────────────────────────────────────────────
        """try:
            import ollama as _ollama
            # Cambiamos a phi3 (mucho más ligero y rápido que llama3)
            # num_ctx: 128 (no necesitamos memoria de contexto para 1 frase)
            # num_predict: 40 (suficiente para 150-200 caracteres)
            response = _ollama.chat(model='phi3', messages=[
                {
                    'role': 'user',
                    'content': f"Descripción técnica corta (1 frase, máximo 150 caracteres) en español para: {product_name}.",
                },
            ], options={
                'num_predict': 40,
                'temperature': 0.1,
                'num_ctx': 128,
                'num_thread': 4, # Ajusta según los cores de tu CPU
            })
            description = response['message']['content'].strip()
            if description:
                description = description.strip('\"').strip("'").strip()
                return {
                    'description': _truncate_to_limit(description, 200),
                    'source': 'Ollama',
                }
        except Exception as ollama_exc:
            _logger.warning(f"Ollama suggest_description error: {ollama_exc}")"""
        

        # ── 1. Wikipedia ES: búsqueda + resumen ───────────────────────────────
        try:
            search_url = 'https://es.wikipedia.org/w/api.php'
            search_params = {
                'action': 'query', 'list': 'search',
                'srsearch': product_name, 'format': 'json', 'srlimit': 1,
            }
            s_resp = _requests.get(search_url, params=search_params,
                                   headers=_HEADERS, timeout=5)
            s_resp.raise_for_status()
            hits = s_resp.json().get('query', {}).get('search', [])
            if hits:
                title = hits[0]['title']
                summary_url = (
                    'https://es.wikipedia.org/api/rest_v1/page/summary/'
                    + urllib.parse.quote(title)
                )
                p_resp = _requests.get(
                    summary_url, headers=_HEADERS, timeout=5)
                if p_resp.status_code == 200:
                    extract = (p_resp.json().get('extract') or '').strip()
                    if extract:
                        cleaned = _clean_extract(extract)
                        return {
                            'description': _truncate_to_limit(cleaned, 200),
                            'source': 'Wikipedia',
                        }
        except Exception as exc:
            _logger.warning(f"Wikipedia suggest_description error: {exc}")

        # ── 2. Fallback: DuckDuckGo Instant Answer ────────────────────────────
        try:
            ddg_url = (
                'https://api.duckduckgo.com/'
                f'?q={urllib.parse.quote(product_name)}'
                '&format=json&no_redirect=1&no_html=1&skip_disambig=1'
            )
            d_resp = _requests.get(ddg_url, headers=_HEADERS, timeout=5)
            d_resp.raise_for_status()
            data = d_resp.json()
            abstract = (data.get('AbstractText') or '').strip()
            if not abstract:
                for topic in (data.get('RelatedTopics') or []):
                    text = (topic.get('Text') or '').strip(
                    ) if isinstance(topic, dict) else ''
                    if text:
                        abstract = text
                        break
            if abstract:
                return {
                    'description': _truncate_to_limit(abstract, 200),
                    'source': 'DuckDuckGo',
                }
        except Exception as exc:
            _logger.warning(
                f"DuckDuckGo suggest_description fallback error: {exc}")

        return {'description': '', 'error': 'No se encontró información para este producto.'}

    # ──────────────────────────────────────────────────────────────────────────
    # Importación masiva por Excel
    # ──────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _normalize_wh_name(name):
        """Replica la normalización que aplica warehouse.product.create():
        quita acentos (NFD→ASCII) y capitaliza la primera letra."""
        cleaned = unicodedata.normalize('NFD', name or '').encode(
            'ascii', 'ignore').decode('ascii')
        return cleaned.capitalize()

    @http.route('/purchase_management/download_mapping_template',
                type='http', auth='user', methods=['GET'])
    def download_mapping_template(self, quotation_id=None, **kw):
        """Genera un Excel con las líneas de la solicitud.
        Si el producto ya existe en el catálogo de bodega, las columnas se
        completan automáticamente con sus datos (color verde).
        Las filas sin coincidencia quedan en blanco y deben completarse manualmente (color amarillo)."""
        try:
            record = request.env['request.quotation'].browse(int(quotation_id))
            if not record.exists():
                return request.make_response(
                    json.dumps({'error': 'Solicitud no encontrada'}),
                    headers=[('Content-Type', 'application/json')], status=404
                )

            product_lines = record.product_line_ids.filtered(
                lambda l: not l.display_type
            ).sorted('sequence')

            # Catálogos para validaciones y búsqueda de coincidencias
            WProduct = request.env['warehouse.product'].sudo()
            WVariant = request.env['warehouse.product.variant'].sudo()
            WAlias = request.env['warehouse.variant.alias'].sudo()
            categories = [c.name for c in request.env['warehouse.product.category'].sudo(
            ).search([], order='name')]
            product_types = [
                t.name for t in request.env['type.product'].sudo().search([], order='name')]

            # Tipo por defecto según la solicitud
            req_type = (record.type_id.name or '').strip().lower()
            default_type = ''
            if req_type == 'productos':
                default_type = next(
                    (t for t in product_types if t.lower() == 'producto'), '')

            wb = openpyxl.Workbook()

            # Nombre amigable usando referencia y tipo de solicitud
            ref = record.reference or str(quotation_id)
            # Pestaña según tipo de solicitud
            type_name = (record.type_id.name or '').strip().lower()
            if 'producto' in type_name:
                tab_name = 'Productos'
            elif 'servicio' in type_name:
                tab_name = 'Servicios'
            else:
                tab_name = 'Items'
            # Archivo: "Productos - REF"
            filename = f'Productos - {ref}.xlsx'

            # ── Hoja: Productos ──
            ws = wb.active
            ws.title = tab_name

            # Estilos
            gray_fill = PatternFill('solid', fgColor='D9D9D9')   # Solo lectura
            yellow_fill = PatternFill(
                'solid', fgColor='FFF2CC')   # Obligatorio, vacío
            # Pre-relleno del catálogo
            green_fill = PatternFill('solid', fgColor='E8F5E9')
            # Variante (producto conocido, falta variante)
            blue_fill = PatternFill('solid', fgColor='E3F2FD')
            header_fill = PatternFill('solid', fgColor='02143f')
            header_font = Font(color='FFFFFF', bold=True, size=10)
            center_align = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'),
            )

            # Cabeceras
            headers = [
                ('Producto solicitado', 40),
                ('Buscar', 10),
                ('Producto bodega', 35),
                ('Categoría', 28),
                ('Tipo', 22),
                ('Variante', 35),
                ('Descripción', 40),
                ('Centro de costo', 30),
            ]
            for col_idx, (title, width) in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_idx)].width = width

            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'A2'

            # Datos con autocompletado y sugerencias
            for row_idx, line in enumerate(product_lines, start=2):
                line_name = line.name or ''

                # Busca primero por alias exacto resuelve producto Y variante de una vez
                alias_variant = self._find_variant_by_alias(WAlias, line_name)

                if alias_variant:
                    # Match completo por alias: todos los campos en verde
                    matched = alias_variant.product_id
                    auto_variant = alias_variant.name
                    existing_variants = [auto_variant]
                    alias_matched = True
                else:
                    alias_matched = False
                    # Busca producto por nombre/fuzzy
                    matched = self._find_catalog_match(WProduct, line_name)
                    existing_variants = []
                    if matched:
                        existing_variants = WVariant.search(
                            [('product_id', '=', matched.id)], order='name'
                        ).mapped('name')
                    # Si el producto tiene una sola variante, se completa automáticamente
                    auto_variant = existing_variants[0] if len(
                        existing_variants) == 1 else ''

                if matched:
                    row_data = [
                        line_name,
                        matched.name,
                        matched.category_id.name if matched.category_id else '',
                        matched.type_product_id.name if matched.type_product_id else default_type,
                        auto_variant,
                        matched.description or '',
                        'Suministro de materiales',
                    ]
                else:
                    row_data = [
                        line_name,
                        '',
                        '',
                        default_type,
                        '',
                        '',
                        'Suministro de materiales',
                    ]

                # Col 1: Producto solicitado (solo lectura, gris)
                c1 = ws.cell(row=row_idx, column=1, value=row_data[0])
                c1.border = thin_border
                c1.alignment = Alignment(vertical='center', wrap_text=False)
                c1.fill = gray_fill

                # Cols 3-8: Datos editables (row_data[1:] → columnas 3..8)
                for i, value in enumerate(row_data[1:], start=3):
                    cell = ws.cell(row=row_idx, column=i, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        vertical='center', wrap_text=False)
                    if matched:
                        if i == 6:  # Variante
                            cell.fill = green_fill if auto_variant else blue_fill
                        else:
                            cell.fill = green_fill
                    else:
                        if i in (4, 5, 8):  # Categoría, Tipo, Centro de costo
                            cell.fill = yellow_fill

                # Comentario solo cuando hay múltiples variantes (sin alias ni auto-fill)
                if existing_variants and not auto_variant:
                    variant_cell = ws.cell(row=row_idx, column=6)
                    hint = 'Variantes existentes en catálogo:\n' + \
                        '\n'.join(f'• {v}' for v in existing_variants)
                    hint += '\n\nEscribe una de las anteriores para reutilizarla\no escribe una nueva para crearla al confirmar.'
                    comment = Comment(hint, 'Sistema')
                    comment.width = 260
                    comment.height = 120 + len(existing_variants) * 16
                    variant_cell.comment = comment

                # Col 2: Buscar en Google (fórmula HYPERLINK nativa de Excel)
                safe_name = line_name.replace('"', "'")
                search_url = 'https://www.google.com/search?q=qué+es+' + \
                    urllib.parse.quote(safe_name, safe='')
                search_cell = ws.cell(row=row_idx, column=2,
                                      value=f'=HYPERLINK("{search_url}","Buscar")')
                search_cell.font = Font(color='1A73E8', size=9)
                search_cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                search_cell.border = thin_border

            data_rows = len(product_lines)
            if data_rows > 0:
                last_row = data_rows + 1

                # Data Validation: Categoría (col E = 5)
                cat_formula = '"' + ','.join(categories) + '"'
                dv_cat = DataValidation(type='list', formula1=cat_formula, allow_blank=True,
                                        showErrorMessage=True,
                                        error='Selecciona una categoría válida',
                                        errorTitle='Categoría inválida')
                ws.add_data_validation(dv_cat)
                dv_cat.add(f'D2:D{last_row}')

                # Data Validation: Tipo (col E = 5)
                type_formula = '"' + ','.join(product_types) + '"'
                dv_type = DataValidation(type='list', formula1=type_formula, allow_blank=True,
                                         showErrorMessage=True,
                                         error='Selecciona un tipo válido',
                                         errorTitle='Tipo inválido')
                ws.add_data_validation(dv_type)
                dv_type.add(f'E2:E{last_row}')

                # Data Validation: Centro de costo (col H = 8)
                dv_cc = DataValidation(
                    type='list',
                    formula1='"Suministro de materiales,Suministro de equipos"',
                    allow_blank=False,
                    showErrorMessage=True,
                    error='Selecciona Suministro de materiales o Suministro de equipos',
                    errorTitle='CC inválido'
                )
                ws.add_data_validation(dv_cc)
                dv_cc.add(f'H2:H{last_row}')

            # Hoja oculta _meta: quotation_id + line_ids en orden para validar el archivo al subirlo y evitar problemas de orden o líneas faltantes.
            ws_meta = wb.create_sheet('_meta')
            ws_meta.sheet_state = 'hidden'
            ws_meta['A1'] = int(quotation_id)
            for i, line in enumerate(product_lines, start=2):
                ws_meta.cell(row=i, column=1, value=line.id)

            # Activar hoja de datos al abrir
            wb.active = ws

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return request.make_response(
                output.read(),
                headers=[
                    ('Content-Type',
                     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition',
                     f'attachment; filename="{filename}"'),
                ]
            )

        except Exception as e:
            _logger.error(
                f"Error al generar plantilla de mapeo: {str(e)}\n{traceback.format_exc()}")
            return request.make_response(
                json.dumps({'error': str(e)}),
                headers=[('Content-Type', 'application/json')], status=500
            )

    def _find_variant_by_alias(self, WAlias, line_name):
        """Busca una variante por alias exacto del nombre de línea (normalizado).
        Retorna el registro warehouse.product.variant o False."""
        if not line_name:
            return False
        normalized = self._normalize_wh_name(line_name)
        alias = WAlias.search([('name', '=', normalized)], limit=1)
        return alias.variant_id if alias else False

    def _find_catalog_match(self, WProduct, line_name):
        """Busca el producto de bodega más probable para una línea de solicitud.

        Estrategias en orden de precisión:
          1. Coincidencia exacta del nombre normalizado completo.
          2. El nombre de catálogo está contenido en el nombre de la línea
             (la línea es más descriptiva, ej. "Diseno arquitectonico" → producto "Diseno").
          3. Alguna palabra significativa del nombre de la línea coincide exactamente
             con algún nombre de producto (búsqueda IN).
          4. Coincidencia ilike con la primera palabra significativa.
        """
        if not line_name:
            return False

        normalized = self._normalize_wh_name(line_name)
        words = [w for w in normalized.split() if len(w) > 2]

        # Coincidencia exacta del nombre completo normalizado
        product = WProduct.search([('name', '=', normalized)], limit=1)
        if product:
            return product

        # Busca productos cuyo nombre esté contenido en el nombre de la línea
        #    (ej. producto "Diseño" dentro de línea "Diseño arquitectónico")
        if words:
            candidates = WProduct.search(
                [('name', 'ilike', words[0])], limit=20
            )
            for p in candidates:
                p_norm = p.name.lower()
                if p_norm in normalized.lower():
                    return p

        # Alguna palabra de la línea coincide exactamente con un nombre de producto
        if words:
            product = WProduct.search(
                [('name', 'in', [w.capitalize() for w in words[:4]])], limit=1
            )
            if product:
                return product

        # ilike con las primeras dos palabras significativas juntas
        if len(words) >= 2:
            product = WProduct.search(
                [('name', 'ilike', words[0])], limit=1
            )
            if product:
                return product

        return False

    @http.route('/purchase_management/import_mapping_excel',
                type='http', auth='user', methods=['POST'], csrf=False)
    def import_mapping_excel(self, quotation_id=None, **kw):
        """Procesa el Excel: crea/busca warehouse.product y variantes (con dedup)
        y devuelve un preview para que el usuario revise antes de confirmar.
        NO cambia el estado de la solicitud — eso lo hace create_and_confirm_mapping.
        Transacción atómica para la creación — si hay errores de validación no se crea nada."""
        def _json_response(data, status=200):
            return request.make_response(
                json.dumps(data),
                headers=[('Content-Type', 'application/json')],
                status=status
            )

        try:
            excel_file = request.httprequest.files.get('excel_file')
            if not excel_file:
                return _json_response({'error': 'No se recibió el archivo'}, 400)

            record = request.env['request.quotation'].browse(int(quotation_id))
            if not record.exists():
                return _json_response({'error': 'Solicitud no encontrada'}, 404)

            if record.state == 'order_created':
                return _json_response({'error': 'Esta solicitud ya fue confirmada. No se puede volver a importar.'}, 400)

            wb = openpyxl.load_workbook(io.BytesIO(
                excel_file.read()), data_only=True)

            # ── Leer _meta ──
            if '_meta' not in wb.sheetnames:
                return _json_response({'error': 'Archivo inválido: falta hoja _meta'}, 400)
            ws_meta = wb['_meta']
            meta_quotation_id = ws_meta.cell(row=1, column=1).value
            if int(meta_quotation_id or 0) != int(quotation_id):
                return _json_response(
                    {'error': 'El archivo no corresponde a esta solicitud'}, 400
                )
            line_ids = []
            for row in ws_meta.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    line_ids.append(int(row[0]))

            data_sheet = next((s for s in wb.sheetnames if s != '_meta'), None)
            if not data_sheet:
                return _json_response({'error': 'Archivo inválido: no se encontró la hoja de datos'}, 400)
            ws = wb[data_sheet]

            # Catálogos válidos para validación
            WCategory = request.env['warehouse.product.category'].sudo()
            WType = request.env['type.product'].sudo()
            CostCenter = request.env['cost.center'].sudo()

            valid_categories = {c.name: c.id for c in WCategory.search([])}
            valid_types = {t.name: t.id for t in WType.search([])}

            # Construir mapa de CCs del proyecto: nombre_analítica_normalizado → cc
            # sudo() garantiza acceso; NFD elimina sensibilidad a tildes/mayúsculas.
            project_cc_map = {}
            if record.project_id:
                for cc in record.project_id.sudo().cost_center_ids:
                    if cc.analytical_account_id:
                        key = unicodedata.normalize('NFD', cc.analytical_account_id.name or '') \
                            .encode('ascii', 'ignore').decode('ascii').lower().strip()
                        project_cc_map[key] = cc

            valid_cc_names = ['Suministro de materiales',
                              'Suministro de equipos']

            # Precio estimado desde cotizaciones adjudicadas
            # Mapa: request_line_id → price_unit de la cotización adjudicada
            price_by_req_line = {}
            awarded_qs = record.quotation_line_ids.filtered(
                lambda q: q.awarded)
            if not awarded_qs:
                awarded_qs = record.quotation_line_ids  # fallback: cualquier cotización
            for q in awarded_qs:
                for pline in q.product_line_ids.filtered(lambda l: not l.display_type):
                    if pline.request_line_id and pline.price_unit > 0:
                        req_id = pline.request_line_id.id
                        if req_id not in price_by_req_line:
                            price_by_req_line[req_id] = pline.price_unit

            # Validar todas las filas ANTES de crear nada ──
            # Si el producto ya existe en BD, los campos categoría/tipo/descripción
            # son opcionales — se toman del registro existente.
            WProduct = request.env['warehouse.product'].sudo()
            errors = []
            parsed_rows = []

            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            # Ignorar filas completamente vacías al final
            data_rows = [r for r in data_rows if any(
                (r[i] is not None and str(r[i]).strip() != '') for i in range(2, 8)
            )]

            if len(data_rows) != len(line_ids):
                return _json_response({
                    'error': (
                        f'El archivo tiene {len(data_rows)} filas de datos pero la solicitud '
                        f'tiene {len(line_ids)} productos. Descarga la plantilla nuevamente.'
                    )
                }, 400)

            # Caché para no repetir búsquedas del mismo producto
            existing_product_cache = {}  # normalized_name -> record or False

            for i, row in enumerate(data_rows):
                row_num = i + 2  # Fila real en Excel (header = 1)
                line_id = line_ids[i]

                product_name_raw = str(row[2] or '').strip()
                category_name = str(row[3] or '').strip()
                type_name = str(row[4] or '').strip()
                variant_name_raw = str(row[5] or '').strip()
                description_raw = str(row[6] or '').strip()
                cc_name = str(row[7] or '').strip()
                estimated_price = price_by_req_line.get(line_id, 1.0)

                # Sentence case: primera letra mayúscula, el resto minúscula
                variant_name = (variant_name_raw[0].upper() + variant_name_raw[1:].lower()
                                if variant_name_raw else '')
                description_cased = (description_raw[0].upper() + description_raw[1:].lower()
                                     if description_raw else '')

                row_errors = []

                # Verificar si el producto ya existe en BD para relajar validaciones
                existing_product = False
                if product_name_raw:
                    normalized = self._normalize_wh_name(product_name_raw)
                    if normalized not in existing_product_cache:
                        existing_product_cache[normalized] = WProduct.search(
                            [('name', '=', normalized)], limit=1
                        ) or False
                    existing_product = existing_product_cache[normalized]

                if not product_name_raw:
                    row_errors.append('Producto bodega es obligatorio')
                else:
                    normalized = self._normalize_wh_name(product_name_raw)
                    if len(normalized.replace(' ', '')) > 25:
                        row_errors.append(
                            f'Nombre "{normalized}" excede 25 caracteres (sin espacios)'
                        )

                if existing_product:
                    # Producto existe: tomar categoría/tipo/descripción del registro en BD
                    # si el usuario no los llenó en el Excel
                    resolved_category_id = existing_product.category_id.id if existing_product.category_id else False
                    resolved_type_id = existing_product.type_product_id.id if existing_product.type_product_id else False
                    resolved_description = description_cased or existing_product.description or ''
                    resolved_category_name = existing_product.category_id.name if existing_product.category_id else ''
                    resolved_type_name = existing_product.type_product_id.name if existing_product.type_product_id else ''
                    # Solo validar categoría/tipo si el usuario los llenó (y son inválidos)
                    if category_name and category_name not in valid_categories:
                        row_errors.append(
                            f'Categoría "{category_name}" no es válida')
                    if type_name and type_name not in valid_types:
                        row_errors.append(f'Tipo "{type_name}" no es válido')
                else:
                    # Producto nuevo: todos los campos son obligatorios
                    resolved_category_id = valid_categories.get(category_name)
                    resolved_type_id = valid_types.get(type_name)
                    resolved_description = description_cased
                    resolved_category_name = category_name
                    resolved_type_name = type_name

                    if not category_name:
                        row_errors.append('Categoría es obligatoria')
                    elif category_name not in valid_categories:
                        row_errors.append(
                            f'Categoría "{category_name}" no es válida')

                    if not type_name:
                        row_errors.append('Tipo es obligatorio')
                    elif type_name not in valid_types:
                        row_errors.append(f'Tipo "{type_name}" no es válido')

                    if not description_cased:
                        row_errors.append('Descripción es obligatoria')
                    else:
                        desc_alnum = re.sub(
                            r'[^a-zA-ZáéíóúÁÉÍÓÚñÑüÜ0-9]', '', description_cased)
                        if len(desc_alnum) > 200:
                            row_errors.append(
                                'Descripción supera 200 caracteres (sin espacios ni caracteres especiales)')

                if not variant_name:
                    row_errors.append('Variante es obligatoria')

                if not cc_name:
                    row_errors.append('Centro de costo es obligatorio')
                elif cc_name not in valid_cc_names:
                    row_errors.append(
                        f'Centro de costo "{cc_name}" no es válido')

                if row_errors:
                    errors.append({
                        'row': row_num,
                        'line_id': line_id,
                        'message': ' | '.join(row_errors),
                    })

                line_rec = request.env['request.product.quotation.line'].sudo().browse(
                    line_id)
                parsed_rows.append({
                    'row_num': row_num,
                    'line_id': line_id,
                    'qty': int(line_rec.qty) if line_rec.exists() else 0,
                    'product_name_raw': product_name_raw,
                    'category_name': resolved_category_name,
                    'type_name': resolved_type_name,
                    'category_id': resolved_category_id,
                    'type_id': resolved_type_id,
                    'variant_name': variant_name,
                    'description': resolved_description,
                    'cc_name': cc_name,
                    'estimated_price': estimated_price,
                })

            if errors:
                rows_result = []
                error_map = {e['line_id']: e['message'] for e in errors}
                Line = request.env['request.product.quotation.line'].sudo()
                for pr in parsed_rows:
                    line_rec = Line.browse(pr['line_id'])
                    rows_result.append({
                        'line_id': pr['line_id'],
                        'line_name': line_rec.name if line_rec.exists() else '',
                        'qty': int(line_rec.qty) if line_rec.exists() else 0,
                        'status': 'error' if pr['line_id'] in error_map else 'pending',
                        'message': error_map.get(pr['line_id'], ''),
                        'variant_id': None,
                        'variant_name': '',
                        'product_name': '',
                        'cost_center_id': None,
                    })
                return _json_response({'mapped': 0, 'errors': errors, 'rows': rows_result})

            # Todo válido: resolver variantes SIN crear — creación diferida al confirmar ──
            # Esto evita que queden registros huérfanos si el usuario cierra el wizard.
            # WProduct ya fue declarado arriba en la sección de validación.
            WVariant = request.env['warehouse.product.variant'].sudo()

            # Reutilizar caché de productos ya buscados durante la validación
            product_cache = existing_product_cache
            rows_result = []
            mappings = []

            for pr in parsed_rows:
                normalized_name = self._normalize_wh_name(
                    pr['product_name_raw'])
                # IDs ya resueltos durante la validación (de BD o del Excel según corresponda)
                category_id = pr['category_id']
                type_id = pr['type_id']

                # Buscar warehouse.product (sin crear)
                if normalized_name not in product_cache:
                    existing = WProduct.search(
                        [('name', '=', normalized_name)], limit=1)
                    product_cache[normalized_name] = existing or False

                product = product_cache[normalized_name]

                # Buscar variante existente (dedup)
                existing_variant = False
                if product:
                    normalized_variant = unicodedata.normalize('NFD', pr['variant_name']).encode(
                        'ascii', 'ignore').decode('ascii').title()
                    existing_variant = WVariant.search([
                        ('product_id', '=', product.id),
                        ('name', '=', normalized_variant),
                    ], limit=1)

                # Resolver centro de costo usando el mapa del proyecto (NFD-normalizado)
                cc_key = unicodedata.normalize('NFD', pr['cc_name'] or '') \
                    .encode('ascii', 'ignore').decode('ascii').lower().strip()
                cc = project_cc_map.get(cc_key) or False

                row_idx = parsed_rows.index(pr)

                if existing_variant:
                    # Variante ya existe en catálogo → vincular directo
                    pending_creation = None
                    variant_id = existing_variant.id
                    variant_name = existing_variant.name
                    product_name = product.name
                    already_existed = True
                    mappings.append({
                        'line_id': pr['line_id'],
                        'variant_id': variant_id,
                        'cost_center_id': cc.id if cc else False,
                    })
                elif product:
                    # Producto existe pero variante no --> creación diferida (solo variante)
                    pending_creation = {
                        'mode': 'variant',
                        'variant_name': pr['variant_name'],
                        'product_id': product.id,
                        'product_name': product.name,
                        'estimated_price': pr['estimated_price'],
                    }
                    variant_id = None
                    variant_name = pr['variant_name']
                    product_name = product.name
                    already_existed = False
                else:
                    # Ni producto ni variante existen --> creación diferida (producto + variante)
                    pending_creation = {
                        'mode': 'product',
                        'product_name': pr['product_name_raw'].strip(),
                        'category_id': category_id,
                        'type_product_id': type_id,
                        'description': pr['description'],
                        'variant_name': pr['variant_name'],
                        'estimated_price': pr['estimated_price'],
                    }
                    variant_id = None
                    variant_name = pr['variant_name']
                    product_name = pr['product_name_raw'].strip()
                    already_existed = False

                rows_result.append({
                    'line_id': pr['line_id'],
                    'line_name': str(data_rows[row_idx][0] or ''),
                    'qty': pr['qty'],
                    'status': 'ok',
                    'message': 'Variante reutilizada del catálogo' if already_existed else '',
                    'already_existed': already_existed,
                    'variant_id': variant_id,
                    'variant_name': variant_name,
                    'product_name': product_name,
                    'pending_creation': pending_creation,
                    'cost_center_id': cc.id if cc else False,
                    'cost_center_name': cc.display_name if cc else '',
                })

            # NO se confirma aquí --> el usuario debe hacer clic en "Confirmar" en el wizard.
            # Las variantes nuevas se crean en create_and_confirm_mapping vía pendingCreation.

            # Marcar cuántas filas comparten el mismo producto nuevo (para mostrar agrupación en UI)
            product_group_counts = {}
            for row in rows_result:
                pc = row.get('pending_creation')
                if pc and pc.get('mode') == 'product':
                    key = (pc.get('product_name') or '').lower()
                    product_group_counts[key] = product_group_counts.get(
                        key, 0) + 1
            for row in rows_result:
                pc = row.get('pending_creation')
                if pc and pc.get('mode') == 'product':
                    key = (pc.get('product_name') or '').lower()
                    row['shared_product_count'] = product_group_counts.get(
                        key, 1)
                else:
                    row['shared_product_count'] = 1

            return _json_response({
                'success': True,
                'mapped': len(mappings),
                'errors': [],
                'rows': rows_result,
                'mappings': mappings,
            })

        except Exception as e:
            _logger.error(
                f"Error al importar Excel de mapeo: {str(e)}\n{traceback.format_exc()}"
            )
            return _json_response({'error': str(e)}, 500)

    # ── Nueva SDC Wizard — Plantilla y carga completa desde Excel ────────────

    @http.route(
        '/purchase_management/nrw_download_template',
        type='http', auth='user', methods=['GET'],
    )
    def nrw_download_template(self, **kw):
        """Genera plantilla Excel de 2 hojas (Solicitud + Productos) para el wizard Nueva SDC."""
        try:
            type_list    = request.env['request.type'].sudo().search([]).mapped('name')
            project_list = request.env['project.management'].sudo().search([]).mapped('name')
            uom_list     = request.env['warehouse.uom'].sudo().search([]).mapped('name')
            file_content = _build_bulk_import_request_template(type_list, project_list, uom_list)
            att = request.env['ir.attachment'].sudo().create({
                'name': 'Plantilla_Nueva_SDC.xlsx',
                'datas': base64.b64encode(file_content),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'type': 'binary',
                'public': True,
            })
            return request.make_json_response({'url': f'/web/content/{att.id}?download=true'})
        except Exception as e:
            _logger.error(f'nrw_download_template error: {e}\n{traceback.format_exc()}')
            return request.make_json_response({'error': str(e)}, status=500)

    @http.route(
        '/purchase_management/nrw_upload_products',
        type='http', auth='user', methods=['POST'], csrf=False,
    )
    def nrw_upload_products(self, **kw):
        """
        Parsea plantilla 2 hojas (Solicitud + Productos).
        Devuelve {header, rows, errors}.
        """
        import datetime as _dt

        try:
            ufile = request.httprequest.files.get('file')
            if not ufile:
                return request.make_json_response({'error': 'No se recibió archivo'}, status=400)

            try:
                wb = openpyxl.load_workbook(io.BytesIO(ufile.read()), data_only=True)
            except Exception:
                return request.make_json_response(
                    {'error': 'Archivo Excel inválido. Asegúrese de cargar un .xlsx.'}, status=400
                )

            if 'Solicitud' not in wb.sheetnames or 'Productos' not in wb.sheetnames:
                return request.make_json_response(
                    {'error': 'El archivo debe tener las hojas "Solicitud" y "Productos". Descargue la plantilla oficial.'},
                    status=400
                )

            errors = []

            # ── Hoja 1: Solicitud ──────────────────────────────────────────
            ws_sol = wb['Solicitud']
            expected_sol = ['Tipo de Solicitud', 'Proyecto', 'Asunto', 'Fecha Límite', 'Motivo']
            actual_sol   = [str(ws_sol.cell(1, c).value or '').strip() for c in range(1, 6)]
            if actual_sol != expected_sol:
                return request.make_json_response({
                    'error': f'Hoja "Solicitud": encabezados incorrectos. Descargue la plantilla oficial.'
                }, status=400)

            row2 = [ws_sol.cell(2, c).value for c in range(1, 6)]
            if not any(row2):
                return request.make_json_response(
                    {'error': 'La hoja "Solicitud" no tiene datos en la fila 2.'}, status=400
                )

            type_raw    = str(row2[0] or '').strip()
            project_raw = str(row2[1] or '').strip()
            subject_raw = str(row2[2] or '').strip()
            deadline_raw = row2[3]
            reason_raw  = str(row2[4] or '').strip()

            if not type_raw:    errors.append('Solicitud: Tipo de Solicitud vacío')
            if not project_raw: errors.append('Solicitud: Proyecto vacío')
            if not subject_raw: errors.append('Solicitud: Asunto vacío')
            if not reason_raw:  errors.append('Solicitud: Motivo vacío')

            # Resolver IDs
            type_rec    = request.env['request.type'].sudo().search(
                [('name', '=ilike', type_raw)], limit=1)
            project_rec = request.env['project.management'].sudo().search(
                [('name', '=ilike', project_raw)], limit=1)

            if type_raw and not type_rec:
                errors.append(f'Solicitud: Tipo "{type_raw}" no encontrado en el sistema')
            if project_raw and not project_rec:
                errors.append(f'Solicitud: Proyecto "{project_raw}" no encontrado en el sistema')

            # Fecha límite
            deadline_str = ''
            if deadline_raw is None or str(deadline_raw).strip() == '':
                errors.append('Solicitud: Fecha Límite vacía')
            else:
                if isinstance(deadline_raw, (_dt.date, _dt.datetime)):
                    deadline_str = deadline_raw.strftime('%Y-%m-%d')
                else:
                    deadline_str = str(deadline_raw).strip()[:10]

            # ── Hoja 2: Productos ──────────────────────────────────────────
            ws_prod = wb['Productos']
            expected_prod = ['Nombre', 'Especificación', 'Cantidad', 'Unidad de Medida']
            actual_prod   = [str(ws_prod.cell(1, c).value or '').strip() for c in range(1, 5)]
            if actual_prod != expected_prod:
                return request.make_json_response({
                    'error': f'Hoja "Productos": encabezados incorrectos. Descargue la plantilla oficial.'
                }, status=400)

            uom_map = {
                r.name.strip().lower(): (r.id, r.name)
                for r in request.env['warehouse.uom'].sudo().search([])
            }

            rows = []
            for row_num in range(2, ws_prod.max_row + 1):
                nombre  = ws_prod.cell(row_num, 1).value
                spec    = ws_prod.cell(row_num, 2).value
                qty_raw = ws_prod.cell(row_num, 3).value
                uom_raw = ws_prod.cell(row_num, 4).value

                if not any([nombre, spec, qty_raw, uom_raw]):
                    continue

                nombre  = str(nombre  or '').strip()
                spec    = str(spec    or '').strip()
                uom_raw = str(uom_raw or '').strip()
                row_err = []

                if not nombre:
                    row_err.append(f'Productos fila {row_num}: nombre vacío')

                qty = None
                if qty_raw is None or str(qty_raw).strip() == '':
                    row_err.append(f'Productos fila {row_num}: cantidad vacía')
                else:
                    try:
                        qty = float(qty_raw)
                        if qty <= 0:
                            row_err.append(f'Productos fila {row_num}: cantidad debe ser > 0')
                    except (ValueError, TypeError):
                        row_err.append(f'Productos fila {row_num}: cantidad "{qty_raw}" inválida')

                uom_id, uom_name = None, uom_raw
                if not uom_raw:
                    row_err.append(f'Productos fila {row_num}: unidad de medida vacía')
                else:
                    match = uom_map.get(uom_raw.lower())
                    if match:
                        uom_id, uom_name = match
                    else:
                        row_err.append(f'Productos fila {row_num}: UdM "{uom_raw}" no encontrada')

                if row_err:
                    errors.extend(row_err)
                else:
                    full_name = nombre if not spec else f'{nombre} — {spec}'
                    rows.append({'name': full_name, 'qty': qty, 'uomId': uom_id, 'uomName': uom_name})

            if not rows and not any('Productos' in e for e in errors):
                errors.append('La hoja "Productos" no contiene datos. Agregue al menos una fila.')

            header = {
                'typeId':       type_rec.id   if type_rec    else None,
                'typeName':     type_rec.name  if type_rec    else type_raw,
                'projectId':    project_rec.id if project_rec else None,
                'projectName':  project_rec.name if project_rec else project_raw,
                'subject':      subject_raw,
                'deadline':     deadline_str,
                'reason':       reason_raw,
            }

            return request.make_json_response({'header': header, 'rows': rows, 'errors': errors})

        except Exception as e:
            _logger.error(f'nrw_upload_products error: {e}\n{traceback.format_exc()}')
            return request.make_json_response({'error': str(e)}, status=500)
