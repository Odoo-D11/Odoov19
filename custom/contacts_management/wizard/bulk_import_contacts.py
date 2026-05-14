# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
from markupsafe import Markup
from ..utils.utils import _build_contacts_import_template
import base64
import io
import openpyxl
import re


class ContactsBulkImportWizard(models.TransientModel):
    _name = 'contacts.bulk.import.wizard'
    _description = 'Wizard para importar contactos desde Excel'
    _transient_max_count = 100
    _transient_max_hours = 24

    """BINARY"""
    file_data = fields.Binary(string='Archivo Excel')
    """CHAR"""
    file_name = fields.Char(string='Nombre del archivo')
    """HTML"""
    validation_html = fields.Html(
        string='Resultado de validación', readonly=True)
    """BOOLEAN"""
    show_instructions = fields.Boolean(
        string='Mostrar instrucciones', default=True, readonly=True)
    show_upload = fields.Boolean(
        string='Mostrar carga', default=False, readonly=True)
    show_validation = fields.Boolean(
        string='Mostrar validación', default=False, readonly=True)
    has_errors = fields.Boolean(
        string='Tiene errores', default=False, readonly=True)

    # =========================================================================
    # NAVEGACION ENTRE PASOS
    # =========================================================================

    def action_download_template(self):
        """Descarga la plantilla Excel con dos hojas: Empresas y Empleados"""
        id_types  = self.env['identification.type'].sudo().search([]).mapped('name')
        countries = self.env['res.country'].sudo().search([]).mapped('name')
        states    = self.env['res.country.state'].sudo().search([]).mapped('name')
        cities    = self.env['res.city'].sudo().search([]).mapped('name')
        titles    = self.env['res.partner.title'].sudo().search([]).mapped('name')
        companies = self.env['res.partner'].sudo().search(
            [('is_business', '=', True)]).mapped('name')
        file_content = _build_contacts_import_template(
            id_types, countries, states, cities, titles, companies)
        file_name = 'Plantilla_Importar_Contactos.xlsx'
        attachment = self.env['ir.attachment'].sudo().create({
            'name': file_name,
            'datas': base64.b64encode(file_content),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'type': 'binary',
            'public': True,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_next_to_upload(self):
        self.ensure_one()
        self.show_instructions = False
        self.show_upload = True
        self.show_validation = False
        return self._reopen_wizard()

    def action_back_to_instructions(self):
        self.ensure_one()
        self.show_instructions = True
        self.show_upload = False
        self.show_validation = False
        self.validation_html = False
        self.has_errors = False
        return self._reopen_wizard()

    def action_back_to_upload(self):
        self.ensure_one()
        self.show_instructions = False
        self.show_upload = True
        self.show_validation = False
        self.validation_html = False
        self.has_errors = False
        return self._reopen_wizard()

    def action_validate_file(self):
        self.ensure_one()
        if not self.file_data:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error'),
                    'message': _('No se ha seleccionado ningún archivo. Por favor adjunte un archivo Excel para continuar.'),
                    'type': 'danger',
                    'sticky': False,
                }
            }
        try:
            empresas, empleados = self._analyze_excel()
            errors = self._validate_contacts(empresas, empleados)
            if errors:
                self.validation_html = self._generate_error_html(errors)
                self.has_errors = True
            else:
                self.validation_html = self._generate_success_html(
                    len(empresas), len(empleados), len(empresas) + len(empleados))
                self.has_errors = False
            self.show_upload = False
            self.show_validation = True
            return self._reopen_wizard()
        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error'),
                    'message': _('Error al procesar el archivo: %s') % str(e),
                    'type': 'danger',
                    'sticky': False,
                }
            }

    def action_process_file(self):
        """Crea los contactos en la base de datos"""
        self.ensure_one()
        try:
            empresas, empleados = self._analyze_excel()
            errors = self._validate_contacts(empresas, empleados)
            if errors:
                raise UserError(
                    _('El archivo contiene errores. Por favor corrija y vuelva a intentar.'))
            created = self._create_contacts(empresas, empleados)
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'res.partner',
                'view_mode': 'list,form',
                'domain': [('id', 'in', [c.id for c in created])],
                'target': 'current',
                'name': _('Contactos Importados'),
            }
        except UserError:
            raise
        except Exception as e:
            raise UserError(_('Error al procesar el archivo: %s') % str(e))

    def _reopen_wizard(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    # =========================================================================
    # ANALISIS DEL EXCEL
    # =========================================================================

    def _analyze_excel(self):
        """
        Lee el archivo Excel y extrae filas de las hojas 'Empresas' y 'Empleados'.
        Retorna (empresas, empleados) — dos listas de dicts.
        Al menos una de las dos hojas debe existir y contener datos.
        """
        self.ensure_one()
        file_data = base64.b64decode(self.file_data)
        file_obj  = io.BytesIO(file_data)
        try:
            workbook = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception:
            raise UserError(
                _('Error al leer el archivo Excel. Asegúrese de que sea un archivo .xlsx válido.'))

        sheet_names = workbook.sheetnames

        # Encabezados esperados por hoja
        empresas_headers = [
            'Nombre',
            'Tipo de Identificación',
            'NIT / Número de Identificación',
            'País',
            'Departamento',
            'Ciudad',
            'Dirección',
            'Teléfono',
            'Correo Electrónico',
        ]
        empleados_headers = [
            'Nombre',
            'Tipo de Identificación',
            'NIT / Número de Identificación',
            'Empresa',
            'País',
            'Departamento',
            'Ciudad',
            'Dirección',
            'Teléfono',
            'Correo Electrónico',
            'Título',
        ]

        def _parse_sheet(wb, sheet_name, expected_headers, row_mapper):
            """
            Parsea una hoja del workbook. Si la hoja no existe devuelve lista vacía.
            Valida encabezados y retorna lista de dicts con los datos de cada fila.
            """
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            actual_headers = [cell.value for cell in ws[1]][:len(expected_headers)]
            if actual_headers != expected_headers:
                raise UserError(
                    _('La hoja "%s" no tiene los encabezados correctos.\n\n'
                      'Esperados: %s\n'
                      'Encontrados: %s\n\n'
                      'Por favor descargue la plantilla correcta.')
                    % (sheet_name,
                       ', '.join(expected_headers),
                       ', '.join([str(h) for h in actual_headers]))
                )
            rows = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                rows.append(row_mapper(row, row_num))
            return rows

        def _row_to_empresa(row, row_num):
            return {
                'sheet':       'Empresas',
                'row_num':     row_num,
                'name':        row[0] if len(row) > 0 else None,
                'id_type':     row[1] if len(row) > 1 else None,
                'vat':         row[2] if len(row) > 2 else None,
                'country':     row[3] if len(row) > 3 else None,
                'state':       row[4] if len(row) > 4 else None,
                'city':        row[5] if len(row) > 5 else None,
                'street':      row[6] if len(row) > 6 else None,
                'phone':       row[7] if len(row) > 7 else None,
                'email':       row[8] if len(row) > 8 else None,
                'is_business': True,
                'is_employee': False,
            }

        def _row_to_empleado(row, row_num):
            return {
                'sheet':       'Empleados',
                'row_num':     row_num,
                'name':        row[0]  if len(row) > 0  else None,
                'id_type':     row[1]  if len(row) > 1  else None,
                'vat':         row[2]  if len(row) > 2  else None,
                'empresa':     row[3]  if len(row) > 3  else None,
                'country':     row[4]  if len(row) > 4  else None,
                'state':       row[5]  if len(row) > 5  else None,
                'city':        row[6]  if len(row) > 6  else None,
                'street':      row[7]  if len(row) > 7  else None,
                'phone':       row[8]  if len(row) > 8  else None,
                'email':       row[9]  if len(row) > 9  else None,
                'title':       row[10] if len(row) > 10 else None,
                'is_business': False,
                'is_employee': True,
            }

        if 'Empresas' not in sheet_names and 'Empleados' not in sheet_names:
            raise UserError(
                _('El archivo Excel no contiene las hojas "Empresas" ni "Empleados". '
                  'Por favor descargue la plantilla correcta.'))

        empresas  = _parse_sheet(workbook, 'Empresas',  empresas_headers,  _row_to_empresa)
        empleados = _parse_sheet(workbook, 'Empleados', empleados_headers, _row_to_empleado)

        if not empresas and not empleados:
            raise UserError(
                _('Las hojas "Empresas" y "Empleados" están vacías. '
                  'Por favor agregue al menos un contacto.'))

        return empresas, empleados

    # =========================================================================
    # VALIDACIONES
    # =========================================================================

    def _validate_contacts(self, empresas, empleados):
        """Orquesta todas las validaciones sobre las dos listas de contactos"""
        errors = []

        # Nombres de empresas del propio batch (resolución same-batch en empleados)
        valid_empresa_names = {
            str(e.get('name', '')).strip().lower()
            for e in empresas
            if e.get('name')
        }

        for empresa in empresas:
            errors.extend(self._validate_empresa_required_fields(empresa))
            errors.extend(self._validate_identification_type(empresa))
            errors.extend(self._validate_nit_format(empresa))
            errors.extend(self._validate_country(empresa))
            errors.extend(self._validate_city_state(empresa))

        for empleado in empleados:
            errors.extend(self._validate_empleado_required_fields(empleado))
            errors.extend(self._validate_identification_type(empleado))
            errors.extend(self._validate_nit_format(empleado))
            errors.extend(self._validate_country(empleado))
            errors.extend(self._validate_city_state(empleado))
            errors.extend(self._validate_title(empleado))
            errors.extend(self._validate_parent_company(empleado, valid_empresa_names))
            errors.extend(self._validate_email_format(empleado))

        # Validaciones cruzadas sobre la lista combinada
        errors.extend(self._validate_duplicate_vat(empresas + empleados))

        return errors

    # -------------------------------------------------------------------------
    # Campos obligatorios por tipo
    # -------------------------------------------------------------------------

    def _validate_empresa_required_fields(self, contact):
        """Valida que los campos obligatorios de una empresa no estén vacíos"""
        errors = []
        sheet  = contact['sheet']
        row    = contact['row_num']
        required = {
            'name':    'Nombre',
            'id_type': 'Tipo de Identificación',
            'vat':     'NIT / Número de Identificación',
            'country': 'País',
            'state':   'Departamento',
            'city':    'Ciudad',
            'street':  'Dirección',
        }
        for field, label in required.items():
            value = contact.get(field)
            if not value or str(value).strip() == '':
                errors.append(
                    Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> Falta completar el campo obligatorio <b class='text-primary'>%s</b>") % (sheet, row, label)
                )
        return errors

    def _validate_empleado_required_fields(self, contact):
        """Valida que los campos obligatorios de un empleado no estén vacíos"""
        errors = []
        sheet  = contact['sheet']
        row    = contact['row_num']
        required = {
            'name':    'Nombre',
            'id_type': 'Tipo de Identificación',
            'vat':     'NIT / Número de Identificación',
            'empresa': 'Empresa',
            'country': 'País',
            'state':   'Departamento',
            'city':    'Ciudad',
            'street':  'Dirección',
            'email':   'Correo Electrónico',
            'title':   'Título',
        }
        for field, label in required.items():
            value = contact.get(field)
            if not value or str(value).strip() == '':
                errors.append(
                    Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> Falta completar el campo obligatorio <b class='text-primary'>%s</b>") % (sheet, row, label)
                )
        return errors

    # -------------------------------------------------------------------------
    # Validaciones de catálogos y formato
    # -------------------------------------------------------------------------

    def _validate_identification_type(self, contact):
        """Verifica que el tipo de identificación exista en el sistema"""
        errors = []
        id_type_val = contact.get('id_type')
        if not id_type_val or str(id_type_val).strip() == '':
            return errors  # Ya lo captura _validate_*_required_fields
        id_type = self.env['identification.type'].search([
            ('name', '=ilike', str(id_type_val).strip())
        ], limit=1)
        if not id_type:
            errors.append(
                Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El tipo de identificación <b class='text-primary'>%s</b> no existe en el sistema.") % (
                    contact['sheet'], contact['row_num'], id_type_val)
            )
        else:
            contact['id_type_id']   = id_type.id
            contact['id_type_code'] = id_type.code
        return errors

    def _validate_nit_format(self, contact):
        """Valida el formato del NIT si el tipo de identificación es NIT"""
        errors = []
        id_type_code = contact.get('id_type_code', '')
        vat          = contact.get('vat')
        if not vat or not id_type_code:
            return errors
        vat_str = str(vat).strip()
        if id_type_code == 'NIT':
            if not (vat_str[:1].isdigit() and vat_str[-1:].isdigit()):
                errors.append(
                    Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El NIT <b class='text-primary'>%s</b> es inválido. Debe comenzar y terminar con un número (ej: 1234.567.890-1).") % (
                        contact['sheet'], contact['row_num'], vat_str)
                )
            elif vat_str.count('.') < 2 or '..' in vat_str:
                errors.append(
                    Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El NIT <b class='text-primary'>%s</b> debe contener al menos dos puntos (.) y no consecutivos (ej: 1234.567.890-1).") % (
                        contact['sheet'], contact['row_num'], vat_str)
                )
            elif vat_str.count('-') > 1 or vat_str.endswith('-'):
                errors.append(
                    Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El NIT <b class='text-primary'>%s</b> solo puede contener un guion (-) y no puede estar al final (ej: 1234.567.890-1).") % (
                        contact['sheet'], contact['row_num'], vat_str)
                )
        return errors

    def _validate_country(self, contact):
        """Verifica que el país exista en res.country"""
        errors = []
        country_val = contact.get('country')
        if not country_val or str(country_val).strip() == '':
            return errors  # Ya lo captura _validate_*_required_fields
        country = self.env['res.country'].search([
            ('name', '=ilike', str(country_val).strip())
        ], limit=1)
        if not country:
            errors.append(
                Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El país <b class='text-primary'>%s</b> no existe en el sistema.") % (
                    contact['sheet'], contact['row_num'], country_val)
            )
        else:
            contact['country_id'] = country.id
        return errors

    def _validate_city_state(self, contact):
        """Valida existencia y consistencia de departamento y ciudad"""
        errors    = []
        sheet     = contact['sheet']
        row       = contact['row_num']
        state_val = contact.get('state')
        city_val  = contact.get('city')

        if state_val and str(state_val).strip():
            state = self.env['res.country.state'].search([
                ('name', '=ilike', str(state_val).strip())
            ], limit=1)
            if not state:
                errors.append(
                    Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El departamento <b class='text-primary'>%s</b> no existe en el sistema.") % (
                        sheet, row, state_val)
                )
            else:
                contact['state_id'] = state.id

        if city_val and str(city_val).strip():
            city_domain = [('name', '=ilike', str(city_val).strip())]
            if contact.get('state_id'):
                city_domain.append(('state_id', '=', contact['state_id']))
            city = self.env['res.city'].search(city_domain, limit=1)
            if not city:
                errors.append(
                    Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> La ciudad <b class='text-primary'>%s</b> no existe en el sistema%s.") % (
                        sheet, row, city_val,
                        Markup(" para el departamento <b class='text-primary'>%s</b>") % state_val if state_val else '')
                )
            else:
                contact['city_id'] = city.id

        return errors

    def _validate_title(self, contact):
        """Verifica que el título exista en res.partner.title; guarda title_id en el dict"""
        errors    = []
        title_val = contact.get('title')
        if not title_val or str(title_val).strip() == '':
            return errors  # Ya lo captura _validate_empleado_required_fields
        title = self.env['res.partner.title'].search([
            ('name', '=ilike', str(title_val).strip())
        ], limit=1)
        if not title:
            errors.append(
                Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El título <b class='text-primary'>%s</b> no existe en el sistema.") % (
                    contact['sheet'], contact['row_num'], title_val)
            )
        else:
            contact['title_id'] = title.id
        return errors

    def _validate_parent_company(self, contact, valid_empresa_names):
        """
        Verifica que la empresa del empleado exista en la BD o en el batch de importación.
        Si se encuentra en la BD, guarda parent_id en el dict.
        Si solo está en el batch, la resolución queda para _create_contacts.
        """
        errors      = []
        empresa_val = contact.get('empresa')
        if not empresa_val or str(empresa_val).strip() == '':
            return errors  # Ya lo captura _validate_empleado_required_fields

        empresa_clean = str(empresa_val).strip()

        # Paso 1: buscar en la base de datos
        empresa = self.env['res.partner'].search([
            ('name', '=ilike', empresa_clean),
            ('is_business', '=', True),
        ], limit=1)
        if empresa:
            contact['parent_id'] = empresa.id
            return []

        # Paso 2: aceptar si el nombre pertenece al batch actual
        if empresa_clean.lower() in valid_empresa_names:
            return []  # parent_id se resolverá en _create_contacts

        errors.append(
            Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> La empresa <b class='text-primary'>%s</b> no existe en el sistema ni en la hoja 'Empresas' de este archivo.") % (
                contact['sheet'], contact['row_num'], empresa_val)
        )
        return errors

    def _validate_email_format(self, contact):
        """Valida el formato del correo electrónico para empleados"""
        errors    = []
        email_val = contact.get('email')
        if not email_val or str(email_val).strip() == '':
            return errors  # Ya lo captura _validate_empleado_required_fields
        email_str = str(email_val).strip()
        pattern   = r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$'
        if not re.match(pattern, email_str):
            errors.append(
                Markup("<b class='text-primary'>Hoja %s, Fila %s:</b> El correo electrónico <b class='text-primary'>%s</b> no tiene un formato válido (ej: usuario@dominio.com).") % (
                    contact['sheet'], contact['row_num'], email_str)
            )
        return errors

    def _validate_duplicate_vat(self, contacts):
        """Detecta NITs/números de identificación duplicados dentro del Excel (ambas hojas)"""
        errors = []
        seen   = {}
        for contact in contacts:
            vat = contact.get('vat')
            if not vat:
                continue
            vat_clean = self._clean_vat(str(vat))
            if vat_clean in seen:
                prev_sheet, prev_row = seen[vat_clean]
                errors.append(
                    Markup("El número de identificación <b class='text-primary'>%s</b> aparece duplicado en el archivo (Hoja <b class='text-primary'>%s</b>, Fila <b class='text-primary'>%s</b> y Hoja <b class='text-primary'>%s</b>, Fila <b class='text-primary'>%s</b>). Los números de identificación deben ser únicos.") % (
                        vat, prev_sheet, prev_row, contact['sheet'], contact['row_num'])
                )
            else:
                seen[vat_clean] = (contact['sheet'], contact['row_num'])
        return errors

    def _clean_vat(self, vat):
        """Elimina puntos, guiones y espacios para comparación"""
        if not vat:
            return ''
        return re.sub(r'[\.\-\s]', '', str(vat)).strip()

    # =========================================================================
    # CREACION DE REGISTROS
    # =========================================================================

    def _create_contacts(self, empresas, empleados):
        """
        Crea los res.partner en orden: primero empresas, luego empleados.
        Los empleados cuya empresa viene del mismo batch se resuelven usando
        el dict created_empresas_by_name.
        """
        created                  = []
        Partner                  = self.env['res.partner']
        created_empresas_by_name = {}

        # 1. Crear empresas
        for empresa in empresas:
            vals = {
                'name':                   str(empresa['name']).strip(),
                'identification_type_id': empresa.get('id_type_id'),
                'vat':                    str(empresa['vat']).strip() if empresa.get('vat') else False,
                'is_business':            True,
                'is_employee':            False,
                'country_id':             empresa.get('country_id'),
                'state_id':               empresa.get('state_id'),
                'city_id':                empresa.get('city_id'),
                'street':                 str(empresa['street']).strip() if empresa.get('street') else False,
                'phone':                  str(empresa['phone']).strip() if empresa.get('phone') else False,
                'email':                  str(empresa['email']).strip().lower() if empresa.get('email') else False,
            }
            record = Partner.sudo().create(vals)
            created.append(record)
            created_empresas_by_name[str(empresa['name']).strip().lower()] = record

        # 2. Crear empleados
        for empleado in empleados:
            parent_id = empleado.get('parent_id')
            if not parent_id:
                # Resolver desde el batch de empresas recién creadas
                batch = created_empresas_by_name.get(
                    str(empleado.get('empresa', '')).strip().lower())
                if batch:
                    parent_id = batch.id

            vals = {
                'name':                   str(empleado['name']).strip(),
                'identification_type_id': empleado.get('id_type_id'),
                'vat':                    str(empleado['vat']).strip() if empleado.get('vat') else False,
                'is_business':            False,
                'is_employee':            True,
                'parent_id':              parent_id or False,
                'country_id':             empleado.get('country_id'),
                'state_id':               empleado.get('state_id'),
                'city_id':                empleado.get('city_id'),
                'street':                 str(empleado['street']).strip() if empleado.get('street') else False,
                'phone':                  str(empleado['phone']).strip() if empleado.get('phone') else False,
                'email':                  str(empleado['email']).strip().lower() if empleado.get('email') else False,
                'title':                  empleado.get('title_id') or False,
            }
            record = Partner.sudo().create(vals)
            created.append(record)

        return created

    # =========================================================================
    # GENERACION DE HTML
    # =========================================================================

    def _generate_error_html(self, errors):
        error_rows = ''
        for error in errors:
            error_rows += f'''
            <div class="d-flex align-items-baseline bg-white p-2 mb-2 rounded border">
                <div class="text-primary mr-2 me-2 text-center" style="min-width: 20px;">
                    <i class="fa fa-times-circle" style="font-size: 0.9rem;"></i>
                </div>
                <div class="flex-grow-1">
                    <span class="text-dark" style="font-size: 0.9rem; line-height: 1.5;">{error}</span>
                </div>
            </div>
            '''

        return Markup(f"""
            <div class="bg-white rounded shadow-sm border p-0 overflow-hidden">
                <div class="row no-gutters">
                    <div class="col-md-4 bg-primary text-white d-flex flex-column align-items-center justify-content-center p-4 text-center">
                        <div class="mb-3">
                            <i class="fa fa-exclamation-triangle fa-4x text-white"></i>
                        </div>
                        <h3 class="font-weight-bold m-0 text-white">ERROR</h3>
                        <p class="mb-0 mt-3 text-white" style="opacity: 0.9; font-size: 0.9rem;">
                            Hemos encontrado inconsistencias en los datos. El archivo no puede ser procesado hasta que se realicen las correcciones.
                        </p>
                    </div>
                    <div class="col-md-8 p-0">
                        <div class="px-4 py-3 border-bottom bg-white d-flex justify-content-between align-items-center">
                            <div>
                                <h5 class="text-dark font-weight-bold m-0">Detalle de validación</h5>
                                <small class="text-muted" style="font-size: 0.85rem;">
                                    Revise los registros listados a continuación.
                                </small>
                            </div>
                            <span class="badge badge-primary px-3 py-2" style="font-size: 0.9rem;">
                                {len(errors)} {'registro' if len(errors) == 1 else 'registros'}
                            </span>
                        </div>
                        <div class="bg-light" style="height: 280px; overflow-y: auto; border-bottom: 1px solid #dee2e6; padding: 16px 32px 16px 16px;">
                            {error_rows}
                        </div>
                        <div class="px-4 py-2 bg-white text-right text-end">
                            <small class="text-muted font-italic">
                                <i class="fa fa-info-circle text-primary mr-1 me-1"></i>
                                Edite el archivo Excel en su equipo y vuelva a cargarlo.
                            </small>
                        </div>
                    </div>
                </div>
            </div>
        """)

    def _generate_success_html(self, total_empresas, total_empleados, total):
        label_total    = 'Contacto'    if total          == 1 else 'Contactos'
        label_empresas = 'Empresa'     if total_empresas == 1 else 'Empresas'
        label_empleados = 'Empleado'   if total_empleados == 1 else 'Empleados'

        return Markup(f"""
            <div class="bg-white rounded shadow-sm border p-0 overflow-hidden">
                <div class="row no-gutters align-items-stretch">
                    <div class="col-md-4 bg-primary text-white d-flex flex-column align-items-center justify-content-center p-4 text-center">
                        <div class="mb-3">
                            <i class="fa fa-check-circle fa-4x text-white"></i>
                        </div>
                        <h4 class="font-weight-bold text-white">¡Todo Listo!</h4>
                        <p class="small mb-0" style="opacity: 0.9;">
                            La información es correcta y está lista para ser importada.
                        </p>
                    </div>
                    <div class="col-md-8 p-4">
                        <h5 class="text-dark text-uppercase font-weight-bold mb-4"
                            style="font-size: 12px; letter-spacing: 1px;">
                            Resumen de Importación
                        </h5>
                        <div class="row text-center mb-4">
                            <div class="col-4 border-right border-end">
                                <h2 class="text-primary font-weight-bold m-0">{total}</h2>
                                <small class="text-muted">{label_total}</small>
                            </div>
                            <div class="col-4 border-right border-end">
                                <h2 class="text-primary font-weight-bold m-0">{total_empresas}</h2>
                                <small class="text-muted">{label_empresas}</small>
                            </div>
                            <div class="col-4">
                                <h2 class="text-primary font-weight-bold m-0">{total_empleados}</h2>
                                <small class="text-muted">{label_empleados}</small>
                            </div>
                        </div>
                        <div class="d-flex align-items-center p-3 rounded" style="background-color: #f8f9fa;">
                            <i class="fa fa-arrow-right text-primary mr-3 me-3 fa-lg"></i>
                            <div>
                                <span class="d-block text-dark font-weight-bold">Siguiente paso:</span>
                                <span class="small text-muted">
                                    Haga clic en el botón "Guardar" para crear los contactos.
                                </span>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        """)
