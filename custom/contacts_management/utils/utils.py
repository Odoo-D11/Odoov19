from difflib import SequenceMatcher
from odoo.exceptions import ValidationError
import unicodedata
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from odoo.addons.utils.models.utils import (  # type: ignore
    convert_first_letter_to_uppercase,
)


def normalize_string(s):
    if isinstance(s, str):
        """ELIMINA TILDES Y CARACTERES ESPECIALES"""
        return ''.join((c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')).lower()


def similar(a, b):
    if not isinstance(a, str) or not isinstance(b, str):
        raise ValidationError(
            'Los valores deben ser cadenas de texto, no números ni listas por favor verifique los datos.')
    a = normalize_string(a)
    b = normalize_string(b)
    return SequenceMatcher(None, a, b).ratio()


def capitalize_first_letter(name):
    return name[:1].upper() + name[1:].lower()


def _build_contacts_import_template(id_types, countries, states, cities, titles, companies):
    """
    Genera el archivo Excel plantilla para importación masiva de contactos.

    Función pura: no accede al ORM. El wizard es responsable de cargar las listas
    y pasarlas como parámetros.

    Args:
        id_types  (list[str]): Nombres de los tipos de identificación (identification.type)
        countries (list[str]): Nombres de los países (res.country)
        states    (list[str]): Nombres de los departamentos (res.country.state)
        cities    (list[str]): Nombres de las ciudades (res.city)
        titles    (list[str]): Nombres de los títulos (res.partner.title)
        companies (list[str]): Nombres de las empresas existentes (res.partner is_business=True)

    Returns:
        bytes: Contenido del archivo .xlsx listo para descargar
    """
    output = io.BytesIO()
    workbook = openpyxl.Workbook()

    # =========================================================================
    # ESTILOS COMPARTIDOS
    # =========================================================================

    corp_dark = '02143f'
    required_blue = 'E8F4FD'
    optional_grey = 'F5F5F5'
    white = 'FFFFFF'

    header_font = Font(bold=True, color=white, size=11)
    header_fill = PatternFill(start_color=corp_dark,
                              end_color=corp_dark,     fill_type='solid')
    required_fill = PatternFill(
        start_color=required_blue, end_color=required_blue, fill_type='solid')
    optional_fill = PatternFill(
        start_color=optional_grey, end_color=optional_grey, fill_type='solid')
    header_align = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(vertical='center', wrap_text=True)
    thin_side = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    # =========================================================================
    # HOJAS OCULTAS DE REFERENCIA (para dropdowns de validación de datos)
    # =========================================================================

    def _add_hidden_sheet(wb, name, values):
        ws = wb.create_sheet(title=name)
        for i, val in enumerate(values, start=1):
            ws.cell(row=i, column=1, value=val)
        ws.sheet_state = 'hidden'

    # Las hojas ocultas deben crearse ANTES de las hojas de datos
    # para que las fórmulas de referencia sean válidas.
    _add_hidden_sheet(workbook, '_Tipos_Identificacion', id_types)
    _add_hidden_sheet(workbook, '_Paises',               countries)
    _add_hidden_sheet(workbook, '_Departamentos',         states)
    _add_hidden_sheet(workbook, '_Ciudades',             cities)
    _add_hidden_sheet(workbook, '_Titulos',              titles)
    _add_hidden_sheet(workbook, '_Empresas',             companies)

    # =========================================================================
    # HELPER: Construye una hoja de datos con estilos y validaciones
    # =========================================================================

    def _make_list_dv(formula, sqref, error_msg='Seleccione un valor de la lista', allow_blank=True):
        dv = DataValidation(
            type='list',
            formula1=formula,
            showDropDown=False,
            allow_blank=allow_blank,
        )
        dv.error = error_msg
        dv.errorTitle = 'Valor no válido'
        dv.showErrorMessage = True
        dv.sqref = sqref
        return dv

    def _build_data_sheet(wb, sheet_name, columns, dropdowns):
        """
        columns  : list of (header_str, is_required: bool, width: int)
        dropdowns: list of (col_letter_range: str, formula: str, allow_blank: bool)
                   e.g. [('B2:B500', \"'_Tipos_Identificacion'!$A$1:$A$10\", True), ...]
        """
        ws = wb.create_sheet(title=sheet_name)

        # Encabezados
        ws.row_dimensions[1].height = 32
        for col_idx, (header, required, width) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Filas de datos 2–500: fondos de color y bordes
        for row_idx in range(2, 501):
            for col_idx, (_, required, _w) in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = required_fill if required else optional_fill
                cell.border = thin_border
                cell.alignment = data_align

        # Congelar encabezado
        ws.freeze_panes = 'A2'

        # Validaciones de datos (dropdowns)
        for sqref, formula, allow_blank in dropdowns:
            ws.add_data_validation(_make_list_dv(
                formula, sqref, allow_blank=allow_blank))

        return ws

    # =========================================================================
    # HOJA "Empresas" — 9 columnas
    # =========================================================================
    #
    # A  Nombre                       obligatorio  libre
    # B  Tipo de Identificación       obligatorio  dropdown _Tipos_Identificacion
    # C  NIT / Número de Id.          obligatorio  libre
    # D  País                         obligatorio  dropdown _Paises
    # E  Departamento                 obligatorio  dropdown _Departamentos
    # F  Ciudad                       obligatorio  dropdown _Ciudades
    # G  Dirección                    obligatorio  libre
    # H  Teléfono                     opcional     libre
    # I  Correo Electrónico           opcional     libre

    empresas_columns = [
        ('Nombre',                         True,  30),
        ('Tipo de Identificación',         True,  25),
        ('NIT / Número de Identificación', True,  32),
        ('País',                           True,  22),
        ('Departamento',                   True,  25),
        ('Ciudad',                         True,  25),
        ('Dirección',                      True,  32),
        ('Teléfono',                       False, 18),
        ('Correo Electrónico',             False, 30),
    ]

    n_id = max(len(id_types),  1)
    n_pais = max(len(countries), 1)
    n_dep = max(len(states),    1)
    n_ciu = max(len(cities),    1)
    n_tit = max(len(titles),    1)
    n_emp = max(len(companies), 1)

    empresas_dropdowns = [
        ('B2:B500', f"'_Tipos_Identificacion'!$A$1:$A${n_id}",  True),
        ('D2:D500', f"'_Paises'!$A$1:$A${n_pais}",              True),
        ('E2:E500', f"'_Departamentos'!$A$1:$A${n_dep}",        True),
        ('F2:F500', f"'_Ciudades'!$A$1:$A${n_ciu}",             True),
    ]

    _build_data_sheet(workbook, 'Empresas',
                      empresas_columns, empresas_dropdowns)

    # =========================================================================
    # HOJA "Empleados" — 11 columnas
    # =========================================================================
    #
    # A  Nombre                       obligatorio  libre
    # B  Tipo de Identificación       obligatorio  dropdown _Tipos_Identificacion
    # C  NIT / Número de Id.          obligatorio  libre
    # D  Empresa                      obligatorio  dropdown _Empresas
    # E  País                         obligatorio  dropdown _Paises
    # F  Departamento                 obligatorio  dropdown _Departamentos
    # G  Ciudad                       obligatorio  dropdown _Ciudades
    # H  Dirección                    obligatorio  libre
    # I  Teléfono                     opcional     libre
    # J  Correo Electrónico           obligatorio  libre
    # K  Título                       obligatorio  dropdown _Titulos

    empleados_columns = [
        ('Nombre',                         True,  30),
        ('Tipo de Identificación',         True,  25),
        ('NIT / Número de Identificación', True,  32),
        ('Empresa',                        True,  30),
        ('País',                           True,  22),
        ('Departamento',                   True,  25),
        ('Ciudad',                         True,  25),
        ('Dirección',                      True,  32),
        ('Teléfono',                       False, 18),
        ('Correo Electrónico',             True,  30),
        ('Título',                         True,  20),
    ]

    empleados_dropdowns = [
        ('B2:B500', f"'_Tipos_Identificacion'!$A$1:$A${n_id}",  True),
        ('D2:D500', f"'_Empresas'!$A$1:$A${n_emp}",             True),
        ('E2:E500', f"'_Paises'!$A$1:$A${n_pais}",              True),
        ('F2:F500', f"'_Departamentos'!$A$1:$A${n_dep}",        True),
        ('G2:G500', f"'_Ciudades'!$A$1:$A${n_ciu}",             True),
        ('K2:K500', f"'_Titulos'!$A$1:$A${n_tit}",              True),
    ]

    _build_data_sheet(workbook, 'Empleados',
                      empleados_columns, empleados_dropdowns)

    # Eliminar la hoja vacía por defecto que crea openpyxl
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    workbook.save(output)
    return output.getvalue()
